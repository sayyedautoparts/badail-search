import io
import os
import re
import sqlite3
import hashlib
import json
import base64
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook
from pydantic import BaseModel

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:  # pragma: no cover
    psycopg = None
    dict_row = None

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
except Exception:  # pragma: no cover
    service_account = None
    build = None
    MediaIoBaseDownload = None


DEFAULT_DB_PATH = "/tmp/search_data.db" if os.getenv("VERCEL") else str(Path(__file__).with_name("search_data.db"))
DB_PATH = Path(os.getenv("APP_DB_PATH", DEFAULT_DB_PATH))
DB_PATH.parent.mkdir(parents=True, exist_ok=True)
SUPABASE_POSTGRES_URL = (
    os.getenv("SUPABASE_DB_URL")
    or os.getenv("SUPABASE_DATABASE_URL")
    or os.getenv("SUPABASE_POOLER_URL")
    or ""
).strip()
NEON_POSTGRES_URL = (os.getenv("NEON_DATABASE_URL") or "").strip()
POSTGRES_DB_URL = (SUPABASE_POSTGRES_URL or NEON_POSTGRES_URL or os.getenv("DATABASE_URL") or "").strip()
USE_POSTGRES = POSTGRES_DB_URL.startswith("postgres")


def detect_db_backend() -> str:
    if not USE_POSTGRES:
        return "sqlite"
    host = POSTGRES_DB_URL.lower()
    if "supabase" in host:
        return "supabase"
    if "neon" in host:
        return "neon"
    return "postgres"


DB_BACKEND = detect_db_backend()
CURRENT_YEAR = datetime.now().year
DIGIT_TRANS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")
ARABIC_LETTER_TRANS = str.maketrans(
    {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ٱ": "ا",
        "ى": "ي",
        "ؤ": "و",
        "ئ": "ي",
        "ة": "ه",
        "ء": "",
        "ـ": "",
    }
)
APP_VERSION = hashlib.sha256(Path(__file__).read_bytes()).hexdigest()[:12]
INGEST_PARSER_VERSION = "5"


class DBConnection:
    def __init__(self, conn: Any, postgres: bool):
        self._conn = conn
        self._postgres = postgres

    def _adapt_sql(self, sql: str) -> str:
        if not self._postgres:
            return sql
        # Keep SQL source code sqlite-friendly and adapt placeholders for postgres.
        return sql.replace("?", "%s")

    def execute(self, sql: str, params: tuple | list = ()):
        return self._conn.execute(self._adapt_sql(sql), params)

    def executemany(self, sql: str, params_seq: list[tuple] | list[list]):
        if not self._postgres:
            return self._conn.executemany(self._adapt_sql(sql), params_seq)
        with self._conn.cursor() as cur:
            cur.executemany(self._adapt_sql(sql), params_seq)
            return cur

    def commit(self) -> None:
        self._conn.commit()

    def close(self) -> None:
        self._conn.close()

    @property
    def postgres(self) -> bool:
        return self._postgres


def get_db() -> DBConnection:
    if USE_POSTGRES:
        if psycopg is None:
            raise RuntimeError("psycopg is not installed; add it to requirements first.")
        conn = psycopg.connect(POSTGRES_DB_URL, row_factory=dict_row)
        return DBConnection(conn, postgres=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return DBConnection(conn, postgres=False)


def init_db() -> None:
    conn = get_db()
    try:
        if conn.postgres:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS uploaded_files (
                    id BIGSERIAL PRIMARY KEY,
                    file_name TEXT UNIQUE NOT NULL,
                    uploaded_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    rows_count INTEGER NOT NULL DEFAULT 0,
                    content_hash TEXT,
                    file_size BIGINT NOT NULL DEFAULT 0,
                    parser_version TEXT NOT NULL DEFAULT ''
                )
                """
            )
        else:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS uploaded_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_name TEXT UNIQUE NOT NULL,
                    uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    rows_count INTEGER NOT NULL DEFAULT 0,
                    content_hash TEXT,
                    file_size INTEGER NOT NULL DEFAULT 0
                )
                """
            )
        # Keep compatibility with older DBs created before these columns existed.
        if conn.postgres:
            columns = {
                row["column_name"]
                for row in conn.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = 'uploaded_files'
                    """
                ).fetchall()
            }
        else:
            columns = {row["name"] for row in conn.execute("PRAGMA table_info(uploaded_files)").fetchall()}
        if "content_hash" not in columns:
            conn.execute("ALTER TABLE uploaded_files ADD COLUMN content_hash TEXT")
        if "file_size" not in columns:
            conn.execute("ALTER TABLE uploaded_files ADD COLUMN file_size INTEGER NOT NULL DEFAULT 0")
        if "parser_version" not in columns:
            conn.execute("ALTER TABLE uploaded_files ADD COLUMN parser_version TEXT NOT NULL DEFAULT ''")
        if conn.postgres:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS products (
                    id BIGSERIAL PRIMARY KEY,
                    item_name TEXT,
                    item_number TEXT,
                    company_number TEXT,
                    original_numbers TEXT,
                    notes TEXT,
                    alternatives TEXT,
                    size_width TEXT,
                    size_diameter TEXT,
                    size_height TEXT,
                    source_file TEXT NOT NULL,
                    source_sheet TEXT NOT NULL
                )
                """
            )
            product_columns = {
                row["column_name"]
                for row in conn.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = 'products'
                    """
                ).fetchall()
            }
        else:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_name TEXT,
                    item_number TEXT,
                    company_number TEXT,
                    original_numbers TEXT,
                    notes TEXT,
                    alternatives TEXT,
                    size_width TEXT,
                    size_diameter TEXT,
                    size_height TEXT,
                    source_file TEXT NOT NULL,
                    source_sheet TEXT NOT NULL
                )
                """
            )
            product_columns = {row["name"] for row in conn.execute("PRAGMA table_info(products)").fetchall()}
        if "company_number" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN company_number TEXT")
        if "notes" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN notes TEXT")
        if "size_width" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN size_width TEXT")
        if "size_diameter" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN size_diameter TEXT")
        if "size_height" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN size_height TEXT")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_item_name ON products(item_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_alternatives ON products(alternatives)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_source_file ON products(source_file)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_company_number ON products(company_number)")
        conn.commit()
    finally:
        conn.close()


def file_fingerprint(file_bytes: bytes) -> tuple[str, int]:
    digest = hashlib.sha256(file_bytes).hexdigest()
    return digest, len(file_bytes)


def is_same_uploaded_file(file_name: str, content_hash: str, file_size: int) -> bool:
    source_key = normalize_source_file_name(file_name)
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT content_hash, file_size, COALESCE(parser_version, '') AS parser_version
            FROM uploaded_files
            WHERE file_name = ?
            """,
            (source_key,),
        ).fetchone()
        if not row:
            return False
        return (
            (row["content_hash"] == content_hash)
            and (int(row["file_size"] or 0) == int(file_size))
            and (str(row["parser_version"] or "") == INGEST_PARSER_VERSION)
        )
    finally:
        conn.close()


def get_sync_meta() -> dict:
    conn = get_db()
    try:
        files = conn.execute(
            """
            SELECT file_name, COALESCE(content_hash, '') AS content_hash, COALESCE(file_size, 0) AS file_size, uploaded_at
            FROM uploaded_files
            ORDER BY file_name
            """
        ).fetchall()
        total_rows = conn.execute("SELECT COUNT(*) AS c FROM products").fetchone()["c"]
        updated_at = conn.execute("SELECT MAX(uploaded_at) AS ts FROM uploaded_files").fetchone()["ts"] or ""

        if files:
            fingerprint_input = "|".join(
                f"{f['file_name']}:{f['content_hash']}:{f['file_size']}:{f['uploaded_at']}" for f in files
            )
        else:
            fingerprint_input = f"rows:{total_rows}:updated:{updated_at}"
        version = hashlib.sha256(fingerprint_input.encode("utf-8")).hexdigest()[:16]
        return {"version": version, "updated_at": updated_at, "total_rows": total_rows, "files_count": len(files)}
    finally:
        conn.close()


def get_sync_rows() -> list[dict]:
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT item_name, item_number, company_number, original_numbers, notes, alternatives,
                   size_width, size_diameter, size_height, source_file, source_sheet
            FROM products
            ORDER BY id
            """
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


class UploadRowsStartIn(BaseModel):
    file_name: str
    content_hash: str
    file_size: int


class UploadRowIn(BaseModel):
    item_name: str = ""
    item_number: str = ""
    company_number: str = ""
    original_numbers: str = ""
    notes: str = ""
    alternatives: str = ""
    size_width: str = ""
    size_diameter: str = ""
    size_height: str = ""
    source_sheet: str = "Sheet1"


class UploadRowsChunkIn(BaseModel):
    file_name: str
    rows: list[UploadRowIn]


class UploadRowsFinishIn(BaseModel):
    file_name: str
    content_hash: str
    file_size: int


class UploadRowsAbortIn(BaseModel):
    file_name: str


def upsert_uploaded_file_meta(file_name: str, rows_count: int, content_hash: str, file_size: int) -> None:
    source_key = normalize_source_file_name(file_name)
    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO uploaded_files (file_name, rows_count, content_hash, file_size, parser_version)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(file_name) DO UPDATE SET
                uploaded_at = CURRENT_TIMESTAMP,
                rows_count = excluded.rows_count,
                content_hash = excluded.content_hash,
                file_size = excluded.file_size,
                parser_version = excluded.parser_version
            """,
            (source_key, rows_count, content_hash, file_size, INGEST_PARSER_VERSION),
        )
        conn.commit()
    finally:
        conn.close()


def normalize_text(text: str) -> str:
    text = str(text).translate(DIGIT_TRANS).translate(ARABIC_LETTER_TRANS).strip().lower()
    text = re.sub(r"[\u064B-\u065F\u0670\u06D6-\u06ED]", "", text)
    text = re.sub(r"[\s\-_]+", " ", text)
    return text


def normalize_text_compact(text: str) -> str:
    n = normalize_text(text)
    return re.sub(r"[\s\-_/]+", "", n)


def _norm_segment_word_set(nseg: str) -> set[str]:
    """كلمات نص مُطبَّع (لتفادي انطباق حرف d داخل كلمة cady)."""
    out: set[str] = set()
    for w in (nseg or "").split():
        w = w.strip(".,|;:()[]")
        if w:
            out.add(w)
    return out


def _token_in_norm_segment(ntok: str, nseg: str) -> bool:
    if not ntok:
        return False
    wset = _norm_segment_word_set(nseg)
    if len(ntok) <= 2:
        return ntok in wset
    if ntok in wset:
        return True
    return ntok in nseg


def _token_appears_in_alternatives_column(ntok: str, alternatives: str) -> bool:
    """التوكن يظهر في عمود البدائل ككلمة مستقلة وليس حرفاً داخل كلمة أخرى."""
    nalt = normalize_text(alternatives)
    if not ntok:
        return False
    wset = _norm_segment_word_set(nalt)
    if len(ntok) <= 2:
        return ntok in wset
    if ntok in wset:
        return True
    return ntok in nalt


def _tokens_matching_alternatives_column(text_tokens: list[str], alternatives: str) -> list[str]:
    """كلمات البحث التي تخص عمود البدائل فقط (تجاهل ما يطابق اسم الصنف أو الملف فقط)."""
    seen: set[str] = set()
    out: list[str] = []
    for t in text_tokens:
        nt = normalize_text(t)
        if not nt or nt in seen:
            continue
        if not _token_appears_in_alternatives_column(nt, alternatives):
            continue
        seen.add(nt)
        out.append(nt)
    return out


def _is_word_char_for_alt_match(ch: str) -> bool:
    """حرف يُعتبر جزءاً من «كلمة» لحدود التطابق (طراز، سعة، إلخ)."""
    return bool(ch) and (ch.isalnum() or ch in "+.")


def _find_whole_word_span(haystack: str, needle: str) -> tuple[int, int] | None:
    """أول ظهور لـ needle في haystack ككلمة كاملة؛ يعيد (بداية، نهاية) على النص الأصلي."""
    if not needle or not haystack:
        return None
    h = haystack.casefold()
    n = needle.casefold()
    pos = 0
    while True:
        idx = h.find(n, pos)
        if idx < 0:
            return None
        end = idx + len(n)
        left_ok = idx == 0 or not _is_word_char_for_alt_match(haystack[idx - 1])
        right_ok = end >= len(haystack) or not _is_word_char_for_alt_match(haystack[end])
        if left_ok and right_ok:
            return (idx, end)
        pos = idx + 1


def _earliest_token_match_span(alternatives: str, tokens: list[str]) -> tuple[int, int] | None:
    """أبكر موضع بين التوكنات (أول كلمة تُرى في الخانة)."""
    best: tuple[int, int] | None = None
    for tok in tokens:
        span = _find_whole_word_span(alternatives, tok)
        if span is None:
            continue
        if best is None or span[0] < best[0]:
            best = span
    return best


def _slice_alternatives_from_first_match_to_slash(alternatives: str, tokens: list[str]) -> str | None:
    """
    من أول ظهور لأحد التوكنات في خانة البدائل حتى أول «/» بعده،
    أو حتى نهاية الخانة إن لم يوجد / (مثل آخر مقطع في الخانة).
    """
    span = _earliest_token_match_span(alternatives, tokens)
    if span is None:
        return None
    st, _ = span
    slash_at = alternatives.find("/", st)
    if slash_at < 0:
        return alternatives[st:].strip()
    return alternatives[st:slash_at].strip()


def tokenize_query(text: str) -> list[str]:
    raw = str(text).translate(DIGIT_TRANS).strip().lower()
    raw = re.sub(r"[،,;/|]+", " ", raw)
    # أقواس حول سعة المحرك مثل (1.6) → 1.6
    raw = re.sub(r"[\(\)]+", " ", raw)
    # دمج سعة المحرك: 1 . 6 أو 1,6 أو 1،6 → 1.6 (السنة بدون نقطة عشرية)
    raw = re.sub(r"(\d{1,2})\s*[.,،]\s*(\d{1,3})(?!\d)", r"\1.\2", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    tokens = [t for t in raw.split(" ") if t]
    # Standalone symbols are separators, not meaningful tokens.
    return [t for t in tokens if t not in {"+", "-", "–", "—"}]


def split_query_display_words(q: str) -> list[str]:
    """كلمات الاستعلام مع الإبقاء على شكل الكتابة (للعرض تحت نتيجة البحث)."""
    raw = str(q).translate(DIGIT_TRANS).strip()
    raw = re.sub(r"[،,;/|]+", " ", raw)
    raw = re.sub(r"[\(\)]+", " ", raw)
    raw = re.sub(r"(\d{1,2})\s*[.,،]\s*(\d{1,3})(?!\d)", r"\1.\2", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    tokens = [t for t in raw.split(" ") if t]
    return [t for t in tokens if t not in {"+", "-", "–", "—"}]


def display_forms_for_text_tokens(text_tokens: list[str], raw_query: str) -> list[str]:
    """يطابق كل توكن بحث مع أول كلمة في الاستعلام الأصلي لنفس التطبيع (للحفاظ على Cady مثلاً)."""
    words = split_query_display_words(raw_query)
    used_word_idx: set[int] = set()
    forms: list[str] = []
    for t in text_tokens:
        nt = normalize_text(t)
        found: str | None = None
        for i, w in enumerate(words):
            if i in used_word_idx:
                continue
            if normalize_text(w) == nt:
                found = w
                used_word_idx.add(i)
                break
        forms.append(found if found is not None else t)
    return forms


def parse_year_token(token: str) -> int | None:
    cleaned = re.sub(r"[^\d]", "", str(token).translate(DIGIT_TRANS))
    if len(cleaned) not in (2, 4):
        return None
    year = int(cleaned)
    if len(cleaned) == 2:
        return 2000 + year if year <= 30 else 1900 + year
    return year


def parse_query_year_token(token: str) -> list[int] | None:
    raw = str(token).translate(DIGIT_TRANS).strip().lower()

    # 1.6 / 2.5 = سعة محرك وليست سنة
    if re.fullmatch(r"\d{1,2}\.\d{1,3}", raw):
        return None

    # +15 or 15+ => 2015 .. current year
    if re.fullmatch(r"(?:\+\d{2,4}|\d{2,4}\+)", raw):
        start_token = raw[1:] if raw.startswith("+") else raw[:-1]
        start = parse_year_token(start_token)
        if start is None:
            return None
        return list(range(start, CURRENT_YEAR + 1))

    # 10-15 => 2010 .. 2015 (range token only)
    match = re.fullmatch(r"(\d{2,4})\s*[-–—]\s*(\d{2,4})", raw)
    if match:
        start = parse_year_token(match.group(1))
        end = parse_year_token(match.group(2))
        if start is None or end is None:
            return None
        low = min(start, end)
        high = max(start, end)
        return list(range(low, high + 1))

    # -07 or 07- => up to 2007 (inclusive).
    if re.fullmatch(r"(?:-\d{2,4}|\d{2,4}-)", raw):
        end_token = raw[1:] if raw.startswith("-") else raw[:-1]
        end = parse_year_token(end_token)
        if end is None:
            return None
        return list(range(1900, end + 1))

    # Single explicit 2/4-digit token is a single year (e.g. 11 => 2011).
    # One-digit tokens (e.g. 2) remain normal text for model numbers.
    if re.fullmatch(r"\d{2,4}", raw):
        year = parse_year_token(raw)
        if year is not None:
            return [year]

    # Other tokens (e.g. 2, 1.6, octavia2) are normal text tokens.
    return None


def is_explicit_year_operator_token(token: str) -> bool:
    raw = str(token).translate(DIGIT_TRANS).strip().lower()
    return bool(
        re.fullmatch(r"(?:\+\d{2,4}|\d{2,4}\+)", raw)
        or re.fullmatch(r"(\d{2,4})\s*[-–—]\s*(\d{2,4})", raw)
        or re.fullmatch(r"(?:-\d{2,4}|\d{2,4}-)", raw)
    )


def year_in_range_text(year: int, text: str) -> bool:
    if not text:
        return False

    # Keep range symbols like "-" and "+" for year parsing.
    normalized = str(text).translate(DIGIT_TRANS).strip().lower()
    normalized = re.sub(r"\s+", " ", normalized)

    # Exact standalone year tokens (e.g. 2011, 11).
    for match in re.finditer(r"(?<!\d)(\d{4})(?!\d)", normalized):
        if int(match.group(1)) == year:
            return True
    for match in re.finditer(r"(?<!\d)(\d{2})(?!\d)", normalized):
        token_year = parse_year_token(match.group(1))
        if token_year == year:
            return True

    # Chained plus markers, e.g. +03,+08 => 03-07 and +08.
    plus_marks: list[tuple[int, int]] = []
    for match in re.finditer(r"(?:\+\s*(\d{2,4})(?!\d)|(?<!\d)(\d{2,4})\s*\+(?!\d))", normalized):
        token = match.group(1) or match.group(2)
        start = parse_year_token(token)
        if start is not None:
            plus_marks.append((match.start(), start))
    if plus_marks:
        plus_marks.sort(key=lambda x: x[0])
        for i, (_, start) in enumerate(plus_marks):
            if i + 1 < len(plus_marks):
                next_start = plus_marks[i + 1][1]
                end = next_start - 1
            else:
                end = CURRENT_YEAR
            if start <= year <= end:
                return True

    # Supports formats like: "lancer +04" or "lancer 04+" => 2004 .. current year.
    for match in re.finditer(r"\+\s*(\d{2,4})(?!\d)", normalized):
        start = parse_year_token(match.group(1))
        if start is not None and start <= year <= CURRENT_YEAR:
            return True
    for match in re.finditer(r"(?<!\d)(\d{2,4})\s*\+(?!\d)", normalized):
        start = parse_year_token(match.group(1))
        if start is not None and start <= year <= CURRENT_YEAR:
            return True

    for match in re.finditer(r"(?<!\d)(\d{2,4})\s*[-–—]\s*(\d{2,4})(?!\d)", normalized):
        start = parse_year_token(match.group(1))
        end = parse_year_token(match.group(2))
        if start is None or end is None:
            continue
        low = min(start, end)
        high = max(start, end)
        if low <= year <= high:
            return True

    # Chained minus markers, e.g. -03,-08 => <=03 and 04-08.
    minus_marks: list[tuple[int, int]] = []
    for match in re.finditer(r"(?:[-–—]\s*(\d{2,4})(?!\d)|(?<!\d)(\d{2,4})\s*[-–—](?!\s*\d))", normalized):
        token = match.group(1) or match.group(2)
        end = parse_year_token(token)
        if end is not None:
            minus_marks.append((match.start(), end))
    if minus_marks:
        minus_marks.sort(key=lambda x: x[0])
        first_end = minus_marks[0][1]
        if year <= first_end:
            return True
        for i in range(1, len(minus_marks)):
            prev_end = minus_marks[i - 1][1]
            cur_end = minus_marks[i][1]
            low = prev_end + 1
            high = cur_end
            if low <= high and low <= year <= high:
                return True

    # Supports open-ended formats like: "-07" or "07-" => up to 2007.
    for match in re.finditer(r"(?<!\d)[-–—]\s*(\d{2,4})(?!\d)", normalized):
        end = parse_year_token(match.group(1))
        if end is not None and year <= end:
            return True
    for match in re.finditer(r"(?<!\d)(\d{2,4})\s*[-–—](?!\s*\d)", normalized):
        end = parse_year_token(match.group(1))
        if end is not None and year <= end:
            return True
    return False


def split_alternative_segments(alternatives: str) -> list[str]:
    if not alternatives:
        return []
    parts = re.split(r"[\/|,\n;]+", alternatives)
    return [p.strip() for p in parts if p and p.strip()]


def split_year_chunks(text: str) -> list[str]:
    if not text:
        return []
    chunks = [
        c.strip()
        for c in re.findall(r"[^/|,\n;]*?(?:\+\s*\d{2,4}|\d{2,4}\s*[-–—]\s*\d{2,4}|\d{2,4}\s*[-–—])", text)
        if c and c.strip()
    ]
    return chunks or [text.strip()]


def year_match_score(segment: str, years: list[int]) -> int:
    if not segment or not years:
        return 0

    normalized = str(segment).translate(DIGIT_TRANS).strip().lower()
    best = 0

    # Exact year token (e.g. 2011 or 11) - highest precision.
    for y in years:
        for match in re.finditer(r"(?<!\d)(\d{4})(?!\d)", normalized):
            if int(match.group(1)) == y:
                best = max(best, 6000)
        for match in re.finditer(r"(?<!\d)(\d{2})(?!\d)", normalized):
            yy = parse_year_token(match.group(1))
            if yy == y:
                best = max(best, 5500)

    # +15 or 15+ => start..current (prefer higher start when matching same year).
    for match in re.finditer(r"(?:\+\s*(\d{2,4})|(?<!\d)(\d{2,4})\s*\+)", normalized):
        token = match.group(1) or match.group(2)
        start = parse_year_token(token)
        if start is None:
            continue
        if all(start <= y <= CURRENT_YEAR for y in years):
            best = max(best, 4000 + start)

    # -07 or 07- => up to end year (prefer tighter upper bound).
    for match in re.finditer(r"(?:[-–—]\s*(\d{2,4})|(?<!\d)(\d{2,4})\s*[-–—](?!\s*\d))", normalized):
        token = match.group(1) or match.group(2)
        end = parse_year_token(token)
        if end is None:
            continue
        if all(y <= end for y in years):
            best = max(best, 3000 + end)

    # 10-15 => closed range (prefer narrower range when matching).
    for match in re.finditer(r"(?<!\d)(\d{2,4})\s*[-–—]\s*(\d{2,4})(?!\d)", normalized):
        start = parse_year_token(match.group(1))
        end = parse_year_token(match.group(2))
        if start is None or end is None:
            continue
        low = min(start, end)
        high = max(start, end)
        if all(low <= y <= high for y in years):
            width = max(1, high - low + 1)
            best = max(best, 5000 - width)

    return best


def extract_matched_alternative(
    alternatives: str,
    text_tokens: list[str],
    years: list[int],
    raw_query: str = "",
) -> str:
    """
    نص العرض تحت النتيجة الزرقاء: من أول ظهور لكلمة من البحث تظهر في خانة البدائل
    حتى أول «/» بعدها، أو حتى آخر الخانة إن لم يوجد / بعدها.
    إن لم ينجح ذلك يُستخدم التطابق بالشرائح كاحتياطي.
    """
    _ = raw_query  # محجوز للتوافق مع الاستدعاءات.
    if not (alternatives or "").strip():
        return ""
    segments = split_alternative_segments(alternatives)
    if not segments:
        return ""

    alt_tokens = _tokens_matching_alternatives_column(text_tokens, alternatives)
    if alt_tokens:
        tokens_use = alt_tokens
    elif years:
        tokens_use = []
    else:
        tokens_use = [normalize_text(t) for t in text_tokens if normalize_text(t)]

    if tokens_use:
        sliced = _slice_alternatives_from_first_match_to_slash(alternatives, tokens_use)
        if sliced:
            if (not years) or any(year_in_range_text(y, sliced) for y in years):
                return sliced

    def seg_matches_tokens(nseg: str) -> bool:
        if not tokens_use:
            return True
        return all(_token_in_norm_segment(tok, nseg) for tok in tokens_use)

    candidates: list[str] = []
    for segment in segments:
        nseg = normalize_text(segment)
        if not seg_matches_tokens(nseg):
            continue
        has_year = (not years) or any(year_in_range_text(y, segment) for y in years)
        if has_year:
            candidates.append(segment)
    if candidates:
        return max(candidates, key=len).strip()

    pool: list[str] = list(segments)
    if years:
        year_pool = [s for s in segments if any(year_in_range_text(y, s) for y in years)]
        if year_pool:
            pool = year_pool

    if tokens_use:
        best_seg = ""
        best_score = -1
        for segment in pool:
            nseg = normalize_text(segment)
            score = sum(1 for tok in tokens_use if _token_in_norm_segment(tok, nseg))
            if score > best_score or (score == best_score and len(segment) > len(best_seg)):
                best_score = score
                best_seg = segment
        if best_score > 0:
            return best_seg.strip()

    if years:
        year_hits = [s for s in segments if any(year_in_range_text(y, s) for y in years)]
        if year_hits:
            return max(year_hits, key=len).strip()
        return ""

    return ""


def row_matches_query(row: sqlite3.Row, text_tokens: list[str], years: list[int]) -> bool:
    """تطابق نصي: كل كلمات البحث يجب أن تظهر في الصف (حقيبة كلمات — ترتيب الكلمات لا يهم)."""
    item_name = row["item_name"] or ""
    alternatives = row["alternatives"] or ""
    source_file = row["source_file"] or ""
    item_norm = normalize_text(item_name)
    alt_norm = normalize_text(alternatives)
    file_norm = normalize_text(source_file)
    combined = normalize_text(f"{item_name} {alternatives} {source_file}")
    combined_compact = normalize_text_compact(f"{item_name} {alternatives} {source_file}")

    for token in text_tokens:
        nt = normalize_text(token)
        nct = normalize_text_compact(token)
        if nt not in combined and (nct and nct not in combined_compact):
            return False

    if years:
        item_tokens = [normalize_text(t) for t in text_tokens if normalize_text(t) in item_norm]
        alt_tokens = [normalize_text(t) for t in text_tokens if normalize_text(t) in alt_norm]
        file_tokens = [normalize_text(t) for t in text_tokens if normalize_text(t) in file_norm]
        item_year_ok = bool(item_tokens) and any(year_in_range_text(y, item_name) for y in years)

        segments = split_alternative_segments(alternatives)
        alt_year_ok = False
        if segments:
            for segment in segments:
                nseg = normalize_text(segment)
                has_alt_tokens = (not alt_tokens) or all(t in nseg for t in alt_tokens)
                has_year = any(year_in_range_text(y, segment) for y in years)
                if has_alt_tokens and has_year:
                    alt_year_ok = True
                    break

        # If query tokens are present in alternatives, year must match alternatives context.
        if alt_tokens:
            if not alt_year_ok:
                return False
        # If tokens are only in source_file context, do not force year match.
        elif file_tokens and not item_tokens:
            pass
        # Otherwise, rely on item-name year context (or alternative year if available).
        elif not (item_year_ok or alt_year_ok):
            return False

    return True


def row_matches_number_query(row: sqlite3.Row, number_tokens: list[str]) -> bool:
    if not number_tokens:
        return True

    number_space = normalize_text(
        f"{row['company_number'] or ''} {row['original_numbers'] or ''} {row['notes'] or ''}"
    )
    return all(normalize_text(t) in number_space for t in number_tokens)


def dedupe_rows_by_normalized_original(rows: list[Any]) -> list[Any]:
    """
    إن تكرر نفس الرقم الأصلي (بعد التطبيع) لأكثر من صف — حتى من ملفات مختلفة —
    يُعتبر صنفاً واحداً ويُعرض صف واحد في جدول النتائج (يُحتفظ بأول صف حسب الترتيب الحالي).
    الصفوف بلا رقم أصلي مُعرَّف لا تُدمج مع بعضها.
    """
    out: list[Any] = []
    seen_orig: set[str] = set()
    for r in rows:
        d = dict(r)
        orig_c = normalize_text_compact(d.get("original_numbers") or "")
        if not orig_c:
            out.append(r)
            continue
        if orig_c in seen_orig:
            continue
        seen_orig.add(orig_c)
        out.append(r)
    return out


def _row_search_relevance_score(row: Any, text_tokens: list[str], years: list[int]) -> tuple[int, int, int, str]:
    """درجة أعلى = الصف أنسب لكلمات البحث الحالية (لاختيار مَن يُبقى عند تكرار الرقم الأصلي)."""
    d = dict(row)
    item = d.get("item_name") or ""
    alt = d.get("alternatives") or ""
    bag = normalize_text(f"{item} {alt}")
    bag_c = normalize_text_compact(f"{item} {alt}")
    item_n = normalize_text(item)
    alt_n = normalize_text(alt)
    alt_c = normalize_text_compact(alt)
    item_c = normalize_text_compact(item)
    score = 0
    for t in text_tokens:
        nt = normalize_text(t)
        if not nt:
            continue
        nct = normalize_text_compact(nt)
        in_bag = nt in bag or (bool(nct) and nct in bag_c)
        if not in_bag:
            continue
        in_alt = nt in alt_n or (bool(nct) and nct in alt_c)
        in_item = nt in item_n or (bool(nct) and nct in item_c)
        if in_alt:
            score += 5
        elif in_item:
            score += 2
    year_bonus = 0
    if years:
        if any(year_in_range_text(y, alt) for y in years):
            year_bonus = 3
        elif any(year_in_range_text(y, item) for y in years):
            year_bonus = 1
    detail = min(len(alt), 5000)
    return (score + year_bonus, year_bonus, detail, item)


def dedupe_rows_by_normalized_original_for_search(
    rows: list[Any],
    text_tokens: list[str],
    years: list[int],
) -> list[Any]:
    """
    دمج تكرار الرقم الأصلي مع اختيار الصف الأكثر ارتباطاً بالاستعلام الحالي
    (كلمات البحث في البدائل تُوزَّن أعلى من الاسم فقط)، حتى لا يختفي صف مثل Octavia
    لأن صفاً آخر بنفس الرقم سبقه أبجدياً (مثل Cady Life).
    """
    winners: dict[str, Any] = {}
    for r in rows:
        d = dict(r)
        oc = normalize_text_compact(d.get("original_numbers") or "")
        if not oc:
            continue
        if oc not in winners:
            winners[oc] = r
        else:
            s_new = _row_search_relevance_score(r, text_tokens, years)
            s_old = _row_search_relevance_score(winners[oc], text_tokens, years)
            if s_new > s_old:
                winners[oc] = r

    out: list[Any] = []
    emitted: set[str] = set()
    for r in rows:
        d = dict(r)
        oc = normalize_text_compact(d.get("original_numbers") or "")
        if not oc:
            out.append(r)
            continue
        if oc in emitted:
            continue
        emitted.add(oc)
        out.append(winners[oc])
    return out


def _numbers_in_normalized_text(text: str) -> list[float]:
    """Extract numeric literals from normalized size text (handles 'قطر 98 مم', '29 x 98', etc.)."""
    out: list[float] = []
    for m in re.finditer(r"\d+(?:\.\d+)?", text or ""):
        try:
            out.append(float(m.group(0)))
        except Exception:
            continue
    return out


def match_size_value(cell_value: str, query_value: str) -> bool:
    q = normalize_text(query_value).replace(",", ".")
    if not q:
        return True
    c = normalize_text(cell_value).replace(",", ".")
    if c == q:
        return True
    # Tolerate different formatting such as "10.0" vs "10"
    try:
        if float(c) == float(q):
            return True
    except Exception:
        pass
    # Query is a plain number but cell has extra text (units, labels, etc.)
    try:
        qf = float(q)
    except Exception:
        return q in c
    for n in _numbers_in_normalized_text(c):
        if n == qf:
            return True
    return False


def match_size_value_driveshaft(cell_value: str, query_value: str) -> bool:
    """
    مطابقة قياسات مصلب الدراي شفط؛ إذا كتب المستخدم رقمًا من رقم واحد (مثل 9 بدل 98)
    يُقبل إذا كان أي رقم في الخلية يبدأ بهذا الرقم (98 تطابق استعلام 9).
    """
    if match_size_value(cell_value, query_value):
        return True
    q = normalize_text(query_value).replace(",", ".")
    if not re.fullmatch(r"\d", q):
        return False
    t = normalize_text(cell_value).replace(",", ".")
    for m in re.finditer(r"\d+(?:\.\d+)?", t):
        s = m.group(0)
        sn = s.lstrip("0") or "0"
        if s.startswith(q) or sn.startswith(q):
            return True
    return False


def normalized_path_matches_driveshaft(sf: str) -> bool:
    """
    True if source_file path looks like a drive-shaft (مصلبات) Excel source.
    Accepts spelling/layout variants often seen in Drive folder names.
    """
    if not sf:
        return False
    if "مصلبة دراي شفط" in sf or "مصلبات دراي شفط" in sf or "مصلب دراي شفط" in sf:
        return True
    # ة -> ه in normalize_text، «مصلب» يغطي مصلبات/مصلبه؛ قد يُكتب «مصلب» بدون تاء
    if "مصلب" not in sf:
        return False
    has_shaft = ("شفط" in sf) or ("shaft" in sf)
    has_drive = any(k in sf for k in ("دراي", "دري", "dry", "drive"))
    # مجلدات باسم «مصلب دراي» بدون «شفط» أو العكس
    return bool(has_shaft or has_drive)


def is_driveshaft_size_query(size_type: str) -> bool:
    st = normalize_text(size_type)
    if not st:
        return False
    if st in (
        "مصلبات دراي شفط",
        "مصلبه دراي شفط",
        "مصلبات",
        "مصلب دراي شفط",
        "مصلب دراي",
    ):
        return True
    if "مصلب" in st and any(x in st for x in ("شفط", "دراي", "دري", "dry", "drive", "shaft")):
        return True
    return False


def file_matches_size_type(source_file: str, size_type: str) -> bool:
    st = normalize_text(size_type)
    if not st:
        return True
    sf = normalize_text(source_file)

    if st == "لبادات":
        return "لبادات" in sf

    if st in ["بيلية", "بيل", "بلي"]:
        return any(key in sf for key in ["نابات + بيل عجل", "بيل مشكلة", "بيـل مشكلة"])

    if is_driveshaft_size_query(size_type):
        return normalized_path_matches_driveshaft(sf)

    return True


def file_matches_size_type_with_fallback(row: Any, source_file: str, size_type: str, sw: str, sd: str) -> bool:
    """مسار الملف أو (مصلبات + قياسان في الاستعلام يطابقان B/C مع تجاهل مسار عند غيبته في source_file)."""
    if not normalize_text(size_type):
        return True
    if file_matches_size_type(source_file, size_type):
        return True
    if not is_driveshaft_size_query(size_type):
        return False
    sf = normalize_text(source_file)
    if "لبادات" in sf:
        return False
    if "بيل مشكلة" in sf or "بيـل مشكلة" in sf or normalize_text("نابات + بيل عجل") in sf:
        return False
    if not (normalize_text(sw) and normalize_text(sd)):
        return False
    rw = (row["size_width"] or "").strip()
    rd = (row["size_diameter"] or "").strip()
    if not rw or not rd:
        return False
    return (match_size_value_driveshaft(rw, sw) and match_size_value_driveshaft(rd, sd)) or (
        match_size_value_driveshaft(rw, sd) and match_size_value_driveshaft(rd, sw)
    )


def row_matches_size_filters(size_type: str, row: Any, sw: str, sd: str, sh: str) -> bool:
    """مطابقة أبعاد البحث؛ لمصلبات الدراي شفط: تبديل B/C مسموح ورقم واحد يطابق بادئة (9≈98)."""
    if is_driveshaft_size_query(size_type):
        wv = row["size_width"] or ""
        dv = row["size_diameter"] or ""
        hv = row["size_height"] or ""
        swq, sdq = normalize_text(sw), normalize_text(sd)
        if swq and sdq:
            dim_ok = (match_size_value_driveshaft(wv, sw) and match_size_value_driveshaft(dv, sd)) or (
                match_size_value_driveshaft(wv, sd) and match_size_value_driveshaft(dv, sw)
            )
        elif swq:
            dim_ok = match_size_value_driveshaft(wv, sw) or match_size_value_driveshaft(dv, sw)
        elif sdq:
            dim_ok = match_size_value_driveshaft(wv, sd) or match_size_value_driveshaft(dv, sd)
        else:
            dim_ok = True
        return dim_ok and match_size_value(hv, sh)
    return (
        match_size_value(row["size_width"] or "", sw)
        and match_size_value(row["size_diameter"] or "", sd)
        and match_size_value(row["size_height"] or "", sh)
    )


def is_size_file(source_file: str) -> bool:
    sf = normalize_text(source_file)
    return (
        ("لبادات" in sf)
        or ("بيل" in sf)
        or ("بيلية" in sf)
        or ("نابات" in sf)
        or normalized_path_matches_driveshaft(sf)
    )


def build_size_label_display(source_file: str, w: str, d: str, h: str) -> str:
    """عرض قياسات بدون «x» الإنجليزية وبدون كلمة Height؛ مصلبات: بُعدان فقط."""
    if not is_size_file(source_file or ""):
        return ""
    sw, sd, sh = (w or "").strip(), (d or "").strip(), (h or "").strip()
    sf = normalize_text(source_file or "")
    if normalized_path_matches_driveshaft(sf) or (is_size_file(source_file or "") and not sh):
        parts = [p for p in (sw, sd) if p]
    else:
        parts = [p for p in (sw, sd, sh) if p]
    return " · ".join(parts) if parts else ""


def preview_row_is_likely_header_row(cells: list[str]) -> bool:
    """يقلل اختيار صف بيانات كأنه صف عناوين (مشكلة ثيرموستات وغيره)."""
    nonempty = [str(c).strip() for c in cells if c and str(c).strip()]
    if len(nonempty) < 2:
        return True
    data_like = 0
    for v in nonempty[:14]:
        if len(v) < 5:
            continue
        dig = sum(1 for ch in v if ch.isdigit())
        if len(v) >= 10 and dig / len(v) > 0.4:
            data_like += 1
        elif len(v) >= 6 and dig / len(v) > 0.55:
            data_like += 1
    return data_like < max(2, len(nonempty) // 3)


def extract_file_search_hints(query: str, tokens: list[str]) -> tuple[list[str], list[str]]:
    qn = normalize_text(query)
    hints: list[str] = []
    consumed_tokens: set[str] = set()

    def consume(*parts: str) -> None:
        for p in parts:
            consumed_tokens.add(normalize_text(p))

    if ("فلتر" in qn or "فلاتر" in qn) and "هواء" in qn:
        hints.extend(["فلتر هواء", "فلاتر هواء"])
        consume("فلتر", "فلاتر", "هواء")
    if ("فلتر" in qn or "فلاتر" in qn) and "زيت" in qn:
        hints.extend(["فلتر زيت", "فلاتر زيت"])
        consume("فلتر", "فلاتر", "زيت")
    if ("فلتر" in qn or "فلاتر" in qn) and ("سولار" in qn or "ديزل" in qn):
        hints.extend(["فلتر سولار", "فلاتر السولار", "فلاتر سولار", "فلتر ديزل", "فلاتر ديزل"])
        consume("فلتر", "فلاتر", "سولار", "ديزل", "السولار")

    cleaned_tokens = [t for t in tokens if normalize_text(t) not in consumed_tokens]
    dedup_hints = list(dict.fromkeys(hints))
    return dedup_hints, cleaned_tokens


def file_search_hint_matches_path(path: str, hint: str) -> bool:
    """
    تطابق تلميح البحث عن الملف مع مسار التخزين.
    يقبل العبارة كاملة في المسار، أو وجود كل كلمات التلميح في المسار منفصلة
    (مثل ملف «فلتر بنزين + سولار» مع تلميح «فلتر سولار»).
    """
    pn = normalize_text(path or "")
    hn = normalize_text(hint or "")
    if not hn:
        return True
    if hn in pn:
        return True
    significant = [w for w in hn.split() if len(w) >= 2]
    if len(significant) < 2:
        return hn in pn

    def word_in_path(w: str) -> bool:
        if w in pn:
            return True
        if w.startswith("ال") and len(w) > 2 and w[2:] in pn:
            return True
        return False

    return all(word_in_path(w) for w in significant)


def file_hints_sql_clause(hints: list[str]) -> tuple[str, list[str]]:
    """
    WHERE فرعي لـ SQLite: لكل تلميح إما LIKE للجملة أو (LIKE لكل كلمة مع AND)
    ليتوافق مع file_search_hint_matches_path.
    """
    or_parts: list[str] = []
    params: list[str] = []
    for h in hints:
        hn = normalize_text(h)
        if not hn:
            continue
        significant = [w for w in hn.split() if len(w) >= 2]
        if len(significant) >= 2:
            and_sql = " AND ".join(["lower(source_file) LIKE ?" for _ in significant])
            or_parts.append(f"({and_sql})")
            params.extend(f"%{w}%" for w in significant)
        else:
            or_parts.append("lower(source_file) LIKE ?")
            params.append(f"%{hn}%")
    if not or_parts:
        return "1=1", []
    return "(" + " OR ".join(or_parts) + ")", params


def get_header_index(header_cells: list[str], keywords: list[str]) -> int | None:
    normalized_headers = [normalize_text(h) for h in header_cells]
    normalized_keywords = [normalize_text(k) for k in keywords]

    for idx, h in enumerate(normalized_headers):
        if h in normalized_keywords:
            return idx
    for idx, h in enumerate(normalized_headers):
        if any(k in h for k in normalized_keywords):
            return idx
    return None


# عمود «رقم الصنف» في الواجهة = رقم الشركات في الملف — يُفصل عن «رقم الصنف» الداخلي في الإكسل
HEADER_COMPANY_NUMBER = [
    "رقم الشركات",
    "رقم الشركة",
    "رقم الشركه",
    "رقم شركة",
    "رقم شركات",
    "ارقام الشركه",
    "أرقام الشركه",
    "ارقام الشركات",
    "أرقام الشركات",
    "ارقام شركات",
    "أرقام شركات",
    "company number",
    "رقم الشركات الموحد",
    "رقم المرجع",
    "المرجع",
    "مرجع الشركة",
    "كود الشركة",
    "كود الصنف",
    "رقم الصنف التجاري",
    "الصنف التجاري",
    "الكود التجاري",
    "كود المنتج",
    "رقم المنتج التجاري",
    "رقم التاجر",
    "رقم المورد",
    "كود التخزين",
    "رقم التعريف",
    "رقم مرجع",
    "part number",
    "reference",
    "stock code",
    "sku",
]
HEADER_ORIGINAL_NUMBERS = [
    "الرقم الاصلي",
    "الرقم الأصلي",
    "رقم اصلي",
    "رقم أصلي",
    "ارقام اصلية",
    "أرقام أصلية",
    "ارقام أصلية",
    "أرقام اصلية",
    "original number",
    "الرقم الاصلي للقطعة",
    "الرقم الأصلي للقطعة",
]
HEADER_ALTERNATIVES = [
    "متشابهات",
    "متشابهاب",
    "المتشابهات",
    "التشابهات",
    "المشابهات",
    "المشابهين",
    "البدائل",
    "بدائل",
    "البديلة",
    "السيارات البديلة",
    "السيارة البديلة",
    "سيارات مشابهة",
    "سيارة مشابهة",
    "alternatives",
    "alternative",
    "سيارات بديلة",
    "السيارات المعادلة",
]
HEADER_ITEM_NAME = [
    "اسم السيارة",
    "اسم الصنف",
    "اسم السيارة/الصنف",
    "النوع",
    "نوع",
    "نوع السيارة",
    "نوع المركبة",
    "نوع المركبه",
    "الصنف",
    "item name",
    "name",
]
# أرقام داخلية / قطعة — بدون رقم الشركات حتى لا يُلتقط عمود الشركات قبل تعيينه صراحة
HEADER_ITEM_NUMBER = [
    "رقم الصنف",
    "رقم القطعة",
    "رقم الصنف الدولي",
    "item number",
    "part number",
]
HEADER_NOTES = ["ملاحظات", "ملاحظة", "notes"]


def _header_has_original_marker(h: str) -> bool:
    return ("اصلي" in h) or ("أصلي" in h) or ("اصليه" in h)


def _header_bad_for_original(h: str) -> bool:
    """لا نعتبر عمود «متشابهات/بدائل» عمودًا للرقم الأصلي إلا إذا ذُكر الأصلي في نفس العنوان."""
    if "متشابه" in h or "تشابه" in h:
        return not _header_has_original_marker(h)
    if "البدائل" in h or "بدائل" in h:
        return not _header_has_original_marker(h)
    return False


def _header_bad_for_alt(h: str) -> bool:
    """لا نلتقط عمود رقم الشركات أو الرقم الأصلي كعمود للبدائل."""
    if "رقم الشركات" in h or "رقم الشركة" in h:
        return True
    if "رقم شركة" in h and "بديل" not in h and "متشابه" not in h and "بدائل" not in h:
        return True
    if _header_has_original_marker(h) and "متشابه" not in h and "بديل" not in h and "بدائل" not in h and "تشابه" not in h:
        return True
    return False


def _header_bad_for_company(h: str) -> bool:
    """رقم الصنف (الواجهة) من عمود رقم الشركات فقط — لا نلتقط عمود النوع/الصنف."""
    if not h:
        return False
    if "رقم الشركات" in h or "رقم الشركة" in h:
        return False
    if "شركات" in h and "رقم" in h:
        return False
    if "شركة" in h and "رقم" in h and "نوع" not in h:
        return False
    if "نوع" in h and "شركة" not in h and "شركات" not in h:
        # «رقم نوع القطعة» وما شابهها قد يكون عمود المرجع التجاري وليس «نوع السيارة» فقط
        if (
            "قطعه" in h
            or "قطعة" in h
            or "القطعه" in h
            or "القطعة" in h
            or "صنف" in h
        ):
            return False
        return True
    if h in ("الصنف", "صنف") or (len(h) <= 12 and "صنف" in h and "رقم" not in h and "شرك" not in h):
        return True
    return False


def find_header_column(
    header_cells: list[str],
    keywords: list[str],
    *,
    used: set[int] | None = None,
    min_substring_len: int = 6,
    reject: Callable[[str], bool] | None = None,
) -> int | None:
    """
    يختار عمودًا واحدًا: مطابقة كاملة للعنوان أولًا، ثم جزئية بشرط طول الكلمة لتفادي أخطاء (مثل احتواء «رقم الصنف» على عنوان طويل خاطئ).
    عند عدة مطابقات: أطول كلمة مفتاحية ثم أقرب عمود لليسار.
    """
    if used is None:
        used = set()
    norm_headers = [normalize_text(h) for h in header_cells]
    sorted_kw = sorted({normalize_text(k) for k in keywords if normalize_text(k)}, key=len, reverse=True)
    if not sorted_kw:
        return None

    for idx, h in enumerate(norm_headers):
        if idx in used or not h:
            continue
        if reject and reject(h):
            continue
        if h in sorted_kw:
            return idx

    best_idx: int | None = None
    best_key: tuple[int, int] | None = None  # (-len(kw), col_index) minimal = longer kw, then left column
    for idx, h in enumerate(norm_headers):
        if idx in used or not h:
            continue
        if reject and reject(h):
            continue
        for kw in sorted_kw:
            if len(kw) < min_substring_len:
                continue
            if kw in h:
                key = (-len(kw), idx)
                if best_key is None or key < best_key:
                    best_key = key
                    best_idx = idx
    return best_idx


# عناوين عمود «رقم الصنف» في الواجهة (بعد normalize_text) — تشمل أشكال الكتابة الشائعة في الإكسل
COMPANY_HEADER_CANONICAL: frozenset[str] = frozenset(
    {
        normalize_text("رقم الشركات"),
        normalize_text("رقم الشركة"),
        normalize_text("رقم شركات"),
        normalize_text("ارقام الشركات"),
        normalize_text("أرقام الشركات"),
        normalize_text("ارقام الشركة"),
        normalize_text("أرقام الشركة"),
    }
)


def normalize_header_cell_for_match(raw: str) -> str:
    """إزالة مسافات غريبة وحروف عرض صفرية قد تمنع مطابقة عنوان «رقم الشركات»."""
    t = normalize_text(str(raw or ""))
    for ch in ("\u200c", "\u200b", "\ufeff", "\xa0"):
        t = t.replace(ch, "")
    return re.sub(r"\s+", " ", t).strip()


def find_company_column_by_exact_header(header_cells: list[str]) -> int | None:
    """
    عمود رقم الصنف = عمود عنوانه يطابق «رقم الشركات» أو مرادفاته بعد التطبيع.
    """
    for idx, raw in enumerate(header_cells):
        h = normalize_header_cell_for_match(raw)
        if not h:
            continue
        if h in COMPANY_HEADER_CANONICAL:
            return idx
        # عنوان طويل قليلاً لكنه واضح (مثل تعليق في الخلية)
        if len(h) <= 32 and "رقم" in h and ("شركات" in h or h.endswith("شركه")):
            if not _header_bad_for_company(h):
                return idx
    return None


def resolve_product_column_indices(header_cells: list[str]) -> dict[str, int | None]:
    used: set[int] = set()

    def assign(
        keywords: list[str],
        min_sub: int,
        reject: Callable[[str], bool] | None = None,
    ) -> int | None:
        idx = find_header_column(
            header_cells, keywords, used=used, min_substring_len=min_sub, reject=reject
        )
        if idx is not None:
            used.add(idx)
        return idx

    # أولوية مطلقة: عمود عنوانه بالضبط «رقم الشركات» في صف العناوين
    company_index = find_company_column_by_exact_header(header_cells)
    if company_index is not None:
        used.add(company_index)
    else:
        company_index = assign(HEADER_COMPANY_NUMBER, 6, reject=_header_bad_for_company)
    original_index = assign(HEADER_ORIGINAL_NUMBERS, 6, reject=_header_bad_for_original)
    alt_index = assign(HEADER_ALTERNATIVES, 5, reject=_header_bad_for_alt)

    if original_index is not None and alt_index == original_index:
        alt_index = None

    item_name_index = assign(HEADER_ITEM_NAME, 4)
    notes_index = assign(HEADER_NOTES, 4)

    item_number_index = find_header_column(
        header_cells, HEADER_ITEM_NUMBER, used=used, min_substring_len=6
    )
    if item_number_index is not None:
        used.add(item_number_index)
    elif company_index is not None:
        item_number_index = company_index
    else:
        item_number_index = 4 if 4 not in used else None
        if item_number_index is None:
            item_number_index = 4

    # مطابقة مرنة لعناوين «رقم/أرقام الشركات» إذا فاتتها القائمة (تاء مربوطة، دمج خلايا، إلخ)
    if company_index is None:
        skip_loose = {x for x in (original_index, alt_index, notes_index) if x is not None}
        for i, raw in enumerate(header_cells):
            if i in skip_loose:
                continue
            h = normalize_text(str(raw or ""))
            if not h:
                continue
            if ("رقم" in h or "كود" in h or "ارقام" in h) and ("شركات" in h or "شركه" in h):
                if _header_bad_for_company(h):
                    continue
                company_index = i
                break

    # غالبًا «رقم الشركات» بجانب «الرقم الأصلي» — جرّب يمين ثم يسار (ملفات ثيرموستات: الأصلي ثم الشركات)
    if company_index is None and original_index is not None:
        blocked = {x for x in (item_name_index, alt_index, original_index, notes_index) if x is not None}
        for delta in (1, -1, 2, -2):
            fb = original_index + delta
            if fb < 0 or fb in blocked:
                continue
            if fb >= len(header_cells):
                continue
            company_index = fb
            break

    return {
        "company_index": company_index,
        "original_index": original_index,
        "alt_index": alt_index,
        "item_name_index": item_name_index,
        "item_number_index": item_number_index,
        "notes_index": notes_index,
    }


def sanitize_alternatives_vs_original(alternatives: str, original_numbers: str) -> str:
    """خانة البدائل تبقى فارغة إن كانت فارغة في الملف؛ ولا نخزّن نصًا مطابقًا للرقم الأصلي."""
    a = (alternatives or "").strip()
    if not a:
        return ""
    o = (original_numbers or "").strip()
    if not o:
        return a
    if normalize_text(a) == normalize_text(o):
        return ""
    # فراغات/أسطر زائدة أو اختلاف بسيط يجعل النص يبدو مثل الأصلي في الواجهة
    if normalize_text_compact(a) == normalize_text_compact(o):
        return ""
    return a


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "none" else text


def forward_fill_header_row(raw_cells: list[object] | tuple[object, ...]) -> list[str]:
    """عند دمج خلايا العنوان أفقيًا تبقى الخلايا التالية فارغة في openpyxl — ننقل عنوان العمود للخلايا الفارغة."""
    cleaned = [clean_cell(v) for v in raw_cells]
    out: list[str] = []
    carry = ""
    for s in cleaned:
        if s:
            carry = s
            out.append(s)
        else:
            out.append(carry)
    return out


def pick_item_name(values: list[str], preferred_idx: int | None) -> str:
    def val(i: int) -> str:
        return values[i].strip() if 0 <= i < len(values) and values[i] else ""

    primary = val(preferred_idx if preferred_idx is not None else -1)
    if primary:
        return primary

    # Prefer early text-like columns if explicit header is missing.
    for i in [1, 0, 2, 3, 4, 5]:
        v = val(i)
        if not v:
            continue
        if re.search(r"[A-Za-z\u0600-\u06FF]", v):
            return v

    # Last fallback: first non-empty cell in early columns.
    for i in range(0, min(10, len(values))):
        v = val(i)
        if v:
            return v
    return ""


def normalize_source_file_name(file_name: str) -> str:
    value = str(file_name or "").strip().replace("\\", "/")
    value = re.sub(r"/+", "/", value).lstrip("./")
    return value or "unnamed.xlsx"


def load_gdrive_service_account_info() -> dict:
    raw_json = (os.getenv("GDRIVE_SERVICE_ACCOUNT_JSON") or "").strip()
    raw_json_b64 = (os.getenv("GDRIVE_SERVICE_ACCOUNT_JSON_B64") or "").strip()
    if raw_json:
        return json.loads(raw_json)
    if raw_json_b64:
        return json.loads(base64.b64decode(raw_json_b64).decode("utf-8"))
    raise RuntimeError("Missing Google service account credentials")


def get_gdrive_service():
    if service_account is None or build is None:
        raise RuntimeError("Google Drive dependencies are not installed")
    info = load_gdrive_service_account_info()
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def is_openpyxl_supported_file_name(file_name: str) -> bool:
    lower = str(file_name or "").lower()
    return lower.endswith(".xlsx") or lower.endswith(".xlsm")


def list_gdrive_excel_files(service, folder_id: str) -> list[dict]:
    excel_mime_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
        "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
        "application/vnd.ms-excel",
    }
    found_files: list[dict] = []
    stack: list[tuple[str, str]] = [(folder_id, "")]
    while stack:
        current_folder_id, current_path = stack.pop()
        page_token = None
        while True:
            resp = (
                service.files()
                .list(
                    q=f"'{current_folder_id}' in parents and trashed = false",
                    fields="nextPageToken, files(id, name, mimeType, modifiedTime, size)",
                    includeItemsFromAllDrives=True,
                    supportsAllDrives=True,
                    pageToken=page_token,
                    pageSize=1000,
                )
                .execute()
            )
            for f in resp.get("files", []):
                name = str(f.get("name") or "").strip()
                if not name:
                    continue
                relative_name = f"{current_path}/{name}" if current_path else name
                mime = str(f.get("mimeType") or "")
                if mime == "application/vnd.google-apps.folder":
                    stack.append((str(f.get("id")), relative_name))
                    continue
                is_excel = mime in excel_mime_types or re.search(r"\.(xlsx|xlsm|xlsb|xls)$", name, re.I)
                if not is_excel:
                    continue
                found_files.append(
                    {
                        "id": str(f.get("id") or ""),
                        "name": name,
                        "relative_name": relative_name,
                        "mimeType": mime,
                        "modifiedTime": str(f.get("modifiedTime") or ""),
                        "size": int(f.get("size") or 0),
                    }
                )
            page_token = resp.get("nextPageToken")
            if not page_token:
                break
    return found_files


def download_gdrive_file_bytes(service, file_id: str) -> bytes:
    if MediaIoBaseDownload is None:
        raise RuntimeError("Google Drive download helper is unavailable")
    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request, chunksize=2 * 1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def process_excel_file(file_bytes: bytes, file_name: str, content_hash: str, file_size: int) -> tuple[int, int]:
    conn = get_db()
    inserted_rows = 0
    unique_item_names: set[str] = set()
    source_key = normalize_source_file_name(file_name)

    try:
        conn.execute("DELETE FROM products WHERE source_file = ?", (source_key,))
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_to_insert = []
                preview_rows = list(ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=40, values_only=True))
                header_candidates = [
                    "اسم الصنف",
                    "اسم السيارة",
                    "رقم الصنف",
                    "رقم الشركة",
                    "رقم الشركات",
                    "الرقم الاصلي",
                    "البدائل",
                    "متشابهات",
                ]
                best_idx = 0
                best_score = -10
                for idx, prow in enumerate(preview_rows):
                    cells = [clean_cell(v) for v in (prow or ())]
                    score = sum(1 for k in header_candidates if get_header_index(cells, [k]) is not None)
                    if not preview_row_is_likely_header_row(cells):
                        score -= 12
                    if score > best_score:
                        best_score = score
                        best_idx = idx
                header_row = preview_rows[best_idx] if preview_rows else ()
                header_cells = forward_fill_header_row(list(header_row or ()))
                data_start_row = best_idx + 2

                cols = resolve_product_column_indices(header_cells)
                company_index = cols["company_index"]
                original_index = cols["original_index"]
                alt_index = cols["alt_index"]
                item_name_index = cols["item_name_index"]
                item_number_index = cols["item_number_index"]
                notes_index = cols["notes_index"]
                if item_name_index is None:
                    item_name_index = 0

                for row in ws.iter_rows(min_row=data_start_row, min_col=1, max_col=40, values_only=True):
                    values = [clean_cell(v) for v in row]

                    def by_idx(idx: int | None, default: str = "") -> str:
                        if idx is None:
                            return default
                        if idx < 0 or idx >= len(values):
                            return default
                        return values[idx]

                    item_name = pick_item_name(values, item_name_index)
                    item_number = by_idx(item_number_index)
                    # رقم الصنف في الواجهة = عمود رقم الشركات؛ إن كانت الخلية فارغة (دمج) نجرّب عمودًا مجاورًا للرقم الأصلي
                    company_number = by_idx(company_index)
                    original_numbers = by_idx(original_index)
                    if (
                        not (company_number or "").strip()
                        and (original_numbers or "").strip()
                        and original_index is not None
                    ):
                        o_c = normalize_text_compact(original_numbers)
                        skip_adj = {
                            i
                            for i in (
                                company_index,
                                alt_index,
                                notes_index,
                                item_name_index,
                                item_number_index,
                            )
                            if i is not None
                        }
                        for delta in (1, -1):
                            j = original_index + delta
                            if j in skip_adj or j < 0 or j >= len(values):
                                continue
                            cand = (values[j] or "").strip() if j < len(values) else ""
                            if not cand or normalize_text_compact(cand) == o_c:
                                continue
                            company_number = cand
                            break
                    notes = by_idx(notes_index)
                    alternatives = by_idx(alt_index) if alt_index is not None else ""
                    alternatives = sanitize_alternatives_vs_original(alternatives, original_numbers)
                    size_width = by_idx(1)
                    size_diameter = by_idx(2)
                    size_height = by_idx(3)

                    if not any(
                        [
                            item_name,
                            item_number,
                            company_number,
                            original_numbers,
                            notes,
                            alternatives,
                            size_width,
                            size_diameter,
                            size_height,
                        ]
                    ):
                        continue

                    rows_to_insert.append(
                        (
                            item_name,
                            item_number,
                            company_number,
                            original_numbers,
                            notes,
                            alternatives,
                            size_width,
                            size_diameter,
                            size_height,
                            source_key,
                            str(sheet_name),
                        )
                    )
                    if item_name:
                        unique_item_names.add(item_name)

                if rows_to_insert:
                    conn.executemany(
                        """
                        INSERT INTO products (
                            item_name, item_number, company_number, original_numbers, notes, alternatives,
                            size_width, size_diameter, size_height, source_file, source_sheet
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        rows_to_insert,
                    )
                    inserted_rows += len(rows_to_insert)
        finally:
            wb.close()

        conn.commit()
        upsert_uploaded_file_meta(source_key, inserted_rows, content_hash, file_size)
        return inserted_rows, len(unique_item_names)
    finally:
        conn.close()


app = FastAPI(title="Advanced Excel Search")
init_db()


@app.get("/", response_class=HTMLResponse)
def home() -> str:
    return """
<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <meta name="theme-color" content="#0b66ff" />
  <meta name="apple-mobile-web-app-capable" content="yes" />
  <meta name="apple-mobile-web-app-status-bar-style" content="default" />
  <meta name="apple-mobile-web-app-title" content="بحث البدائل" />
  <link rel="manifest" href="/manifest.webmanifest" />
  <title>بحث البدائل</title>
  <style>
    body { font-family: Arial, sans-serif; background: #f6f7fb; margin: 0; }
    .wrap { max-width: 760px; margin: 18px auto; padding: 0 12px; }
    .card { background: #fff; border-radius: 12px; padding: 16px; box-shadow: 0 2px 10px rgba(0,0,0,.06); margin-bottom: 16px; }
    input, select, button { width: 100%; padding: 10px; margin: 6px 0; border: 1px solid #ddd; border-radius: 8px; box-sizing: border-box; font-size: 16px; }
    button { background: #0b66ff; color: #fff; border: none; cursor: pointer; }
    button:hover { opacity: 0.95; }
    table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 13px; table-layout: fixed; }
    th, td { border: 1px solid #eee; padding: 6px; text-align: right; vertical-align: top; word-wrap: break-word; }
    #resultsTable th:first-child, #resultsTable td:first-child { width: 14%; font-size: 12px; color: #444; }
    th { background: #fafafa; }
    .muted { color: #666; font-size: 13px; }
    .pill { display: inline-block; background: #eef4ff; color: #2d5fff; padding: 7px 10px; border-radius: 10px; margin: 4px 4px 0 0; font-size: 12px; text-align: right; }
    .pill-btn { border: none; cursor: pointer; min-width: 180px; }
    .quick-name { display: inline; font-weight: 700; }
    .quick-alt { display: block; margin-top: 3px; color: #4e5f8f; }
    .quick-size { display: block; margin-top: 3px; color: #3e4d76; font-size: 12px; }
    .row-focus { background: #fff7d6; transition: background-color 0.25s ease; }
    .upload-label { display: block; margin-top: 8px; font-size: 12px; color: #334; font-weight: 700; }
    .upload-actions { display: flex; gap: 8px; margin-top: 8px; }
    .upload-actions button { flex: 1; }
    .btn-secondary { background: #edf0f8; color: #223; }
    .drop-zone { margin-top: 8px; border: 2px dashed #b8c4e6; border-radius: 10px; padding: 12px; text-align: center; color: #445; background: #f8faff; }
    .drop-zone.active { border-color: #2d5fff; background: #eef3ff; }
    .progress-wrap { margin-top: 8px; height: 10px; background: #e8ecf5; border-radius: 999px; overflow: hidden; }
    .progress-bar { height: 100%; width: 0%; background: #2d5fff; transition: width 0.2s ease; }
    .tabs { display: flex; gap: 8px; margin: 8px 0 10px; }
    .tab-btn { flex: 1; background: #edf0f8; color: #223; }
    .tab-btn.active { background: #0b66ff; color: #fff; }
    .search-panel { display: none; }
    .search-panel.active { display: block; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h2>رفع ملفات Excel</h2>
      <p class="muted">اختَر ملفات Excel و/أو فولدرات تحتوي Excel بنفس الوقت، ثم ارفع دفعة واحدة.</p>
      <button type="button" class="btn-secondary" onclick="pickFolderDeep()">اختيار مجلد رئيسي مع كل المجلدات الفرعية</button>
      <div class="upload-actions">
        <button type="button" class="btn-secondary" onclick="clearSelectedFiles()">تفريغ الاختيار</button>
        <button id="uploadBtn" type="button" onclick="uploadFiles()">رفع الملفات</button>
      </div>
      <div id="selectionInfo" class="muted">الملفات المحددة: 0</div>
      <div id="progressInfo" class="muted"></div>
      <div class="progress-wrap"><div id="progressBar" class="progress-bar"></div></div>
      <div id="uploadResult" class="muted"></div>
      <div class="upload-actions">
        <button type="button" onclick="syncSnapshotFromServer()">مزامنة للهاتف (Offline)</button>
        <button type="button" class="btn-secondary" onclick="checkSnapshotUpdate(true, true)">فحص تحديث البيانات</button>
      </div>
      <div id="syncInfo" class="muted"></div>
    </div>

    <div class="card">
      <h2>البحث</h2>
      <div class="tabs">
        <button id="tabMainBtn" type="button" class="tab-btn active" onclick="setSearchTab('main')">البحث الرئيسي</button>
        <button id="tabSizeBtn" type="button" class="tab-btn" onclick="setSearchTab('size')">بحث القياسات</button>
      </div>

      <div id="mainSearchPanel" class="search-panel active">
        <p class="muted">البحث الرئيسي داخل: اسم الصنف + المتشابهات (البدائل)</p>
        <input id="queryInput" placeholder="بحث رئيسي (اسم الصنف / المتشابهات)..." />
        <input id="numberQueryInput" placeholder="بحث الأرقام (الملاحظات / الرقم الأصلي / رقم الشركة)..." />
        <button onclick="searchNow('main')">بحث</button>
      </div>

      <div id="sizeSearchPanel" class="search-panel">
        <p class="muted" style="margin-top:10px;">بحث القياسات (لبادات / بيل / مصلبات دراي شفط)</p>
        <select id="sizeTypeInput">
          <option value="">اختر النوع (اختياري)</option>
          <option value="لبادات">لبادات</option>
          <option value="بيل">بيل</option>
          <option value="مصلبات دراي شفط">مصلبات دراي شفط</option>
        </select>
        <input id="sizeWidthInput" placeholder="العرض (B)" />
        <input id="sizeDiameterInput" placeholder="القطر (C)" />
        <input id="sizeHeightInput" placeholder="الارتفاع (D)" />
        <button onclick="searchNow('size')">بحث القياسات</button>
      </div>

      <div id="stats" class="muted"></div>
      <div id="names"></div>
      <div style="overflow:auto;">
        <table id="resultsTable" style="display:none;">
          <thead>
            <tr>
              <th>اسم الملف</th>
              <th>اسم الصنف</th>
              <th>رقم الصنف</th>
              <th>السيارات البديلة</th>
            </tr>
          </thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>

  <script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
  <script>
    let autoRefreshTimer = null;
    let searchDebounceTimer = null;
    const selectedFilesMap = new Map();
    let selectedFileSerial = 0;
    const CLIENT_APP_VERSION = '{APP_VERSION}';
    let activeSearchTab = 'main';
    const OFFLINE_DB_NAME = 'badail_offline_db';
    const OFFLINE_DB_STORE = 'snapshots';
    const OFFLINE_DB_KEY = 'latest';
    let localSnapshot = { version: '', updated_at: '', total_rows: 0, rows: [] };
    let lastNotifiedVersion = '';

    function setSyncInfo(message) {
      document.getElementById('syncInfo').innerText = message || '';
    }

    function setSearchTab(tab) {
      activeSearchTab = (tab === 'size') ? 'size' : 'main';
      localStorage.setItem('activeSearchTab', activeSearchTab);
      document.getElementById('tabMainBtn').classList.toggle('active', activeSearchTab === 'main');
      document.getElementById('tabSizeBtn').classList.toggle('active', activeSearchTab === 'size');
      document.getElementById('mainSearchPanel').classList.toggle('active', activeSearchTab === 'main');
      document.getElementById('sizeSearchPanel').classList.toggle('active', activeSearchTab === 'size');
    }

    function registerServiceWorker() {
      if (!('serviceWorker' in navigator)) return;
      navigator.serviceWorker.register('/sw.js').then((reg) => {
        reg.update().catch(() => {});
        reg.addEventListener('updatefound', () => {
          const newWorker = reg.installing;
          if (!newWorker) return;
          newWorker.addEventListener('statechange', () => {
            if (newWorker.state === 'installed' && navigator.serviceWorker.controller) {
              setSyncInfo('يوجد تحديث للتطبيق. أغلقه وافتحه من جديد عند رغبتك.');
            }
          });
        });
      }).catch(() => {
        // ignore service-worker registration errors
      });
    }

    async function ensureLatestClientVersion() {
      try {
        const res = await fetch(`/health?ts=${Date.now()}`, { cache: 'no-store' });
        if (!res.ok) return;
        const data = await res.json();
        const serverVersion = String(data.app_version || '').trim();
        if (!serverVersion) return;

        const localVersion = localStorage.getItem('app_version') || '';
        if (!localVersion) {
          localStorage.setItem('app_version', serverVersion);
          return;
        }

        if (localVersion !== serverVersion) {
          localStorage.setItem('app_version', serverVersion);
          if ('serviceWorker' in navigator) {
            try {
              const regs = await navigator.serviceWorker.getRegistrations();
              await Promise.all(regs.map(r => r.update()));
            } catch (_) {}
          }
          if (!window.location.search.includes(`av=${encodeURIComponent(serverVersion)}`)) {
            const next = `/?av=${encodeURIComponent(serverVersion)}&ts=${Date.now()}`;
            window.location.replace(next);
          }
        }
      } catch (_) {
        // ignore version-check errors
      }
    }

    function openOfflineDb() {
      return new Promise((resolve, reject) => {
        const req = indexedDB.open(OFFLINE_DB_NAME, 1);
        req.onupgradeneeded = () => {
          const db = req.result;
          if (!db.objectStoreNames.contains(OFFLINE_DB_STORE)) {
            db.createObjectStore(OFFLINE_DB_STORE);
          }
        };
        req.onsuccess = () => resolve(req.result);
        req.onerror = () => reject(req.error);
      });
    }

    async function loadLocalSnapshot() {
      try {
        const db = await openOfflineDb();
        const snapshot = await new Promise((resolve, reject) => {
          const tx = db.transaction(OFFLINE_DB_STORE, 'readonly');
          const req = tx.objectStore(OFFLINE_DB_STORE).get(OFFLINE_DB_KEY);
          req.onsuccess = () => resolve(req.result || null);
          req.onerror = () => reject(req.error);
        });
        db.close();
        if (snapshot && Array.isArray(snapshot.rows)) {
          localSnapshot = snapshot;
          setSyncInfo(`نسخة محلية متوفرة: ${snapshot.total_rows || snapshot.rows.length} صف (إصدار ${snapshot.version || '-'})`);
        } else {
          setSyncInfo('لا توجد نسخة محلية بعد. اعمل مزامنة للهاتف.');
        }
      } catch (_) {
        setSyncInfo('تعذر تحميل النسخة المحلية من الجهاز.');
      }
    }

    async function saveLocalSnapshot(snapshot) {
      const payload = {
        version: snapshot.version || '',
        updated_at: snapshot.updated_at || '',
        total_rows: Number(snapshot.total_rows || (snapshot.rows || []).length),
        rows: Array.isArray(snapshot.rows) ? snapshot.rows : []
      };
      const db = await openOfflineDb();
      await new Promise((resolve, reject) => {
        const tx = db.transaction(OFFLINE_DB_STORE, 'readwrite');
        tx.objectStore(OFFLINE_DB_STORE).put(payload, OFFLINE_DB_KEY);
        tx.oncomplete = () => resolve();
        tx.onerror = () => reject(tx.error);
      });
      db.close();
      localSnapshot = payload;
      setSyncInfo(`تمت مزامنة النسخة المحلية: ${payload.total_rows} صف (إصدار ${payload.version || '-'})`);
    }

    function excelOnly(files) {
      return Array.from(files).filter(f => /\\.(xlsx|xlsm|xls|xlsb)$/i.test(f.name));
    }

    function sourceKeyForFile(file) {
      return String(file.webkitRelativePath || file.relativePath || file.name || '')
        .replace(/\\\\/g, '/')
        .replace(/^\\/+/, '')
        .trim() || String(file.name || 'unnamed.xlsx');
    }

    function fileKey(file) {
      // Keep every file entry unique to avoid accidental overwrite
      // when different files share same name/size/date metadata.
      return `${sourceKeyForFile(file)}::${file.size}::${file.lastModified}::${selectedFileSerial++}`;
    }

    function updateSelectionInfo() {
      const count = selectedFilesMap.size;
      document.getElementById('selectionInfo').innerText = `الملفات المحددة: ${count}`;
    }

    function appendFilesToSelection(files) {
      excelOnly(files).forEach(file => {
        selectedFilesMap.set(fileKey(file), file);
      });
      updateSelectionInfo();
    }

    function clearSelectedFiles() {
      selectedFilesMap.clear();
      updateSelectionInfo();
      setUploadProgress(0, '');
      document.getElementById('uploadResult').innerText = 'تم تفريغ الاختيارات.';
    }

    function setUploadProgress(percent, message) {
      const p = Math.max(0, Math.min(100, percent));
      document.getElementById('progressBar').style.width = `${p}%`;
      document.getElementById('progressInfo').innerText = message || '';
    }

    function readFileFromEntry(entry) {
      return new Promise((resolve) => {
        entry.file((file) => {
          try {
            file.relativePath = String(entry.fullPath || file.webkitRelativePath || file.name || '').replace(/^\\/+/, '');
          } catch (_) {}
          resolve([file]);
        }, () => resolve([]));
      });
    }

    function readEntriesOnce(reader) {
      return new Promise((resolve) => {
        reader.readEntries((entries) => resolve(entries || []), () => resolve([]));
      });
    }

    async function collectFilesFromEntry(entry) {
      if (!entry) return [];
      if (entry.isFile) {
        return await readFileFromEntry(entry);
      }
      if (!entry.isDirectory) return [];

      const reader = entry.createReader();
      const allEntries = [];
      while (true) {
        const batch = await readEntriesOnce(reader);
        if (!batch.length) break;
        allEntries.push(...batch);
      }

      const collected = [];
      for (const child of allEntries) {
        const files = await collectFilesFromEntry(child);
        collected.push(...files);
      }
      return collected;
    }

    async function collectFilesFromDirectoryHandle(handle, parentPath = '') {
      const files = [];
      for await (const [name, child] of handle.entries()) {
        const nextPath = parentPath ? `${parentPath}/${name}` : name;
        if (child.kind === 'file') {
          const file = await child.getFile();
          try { file.relativePath = nextPath; } catch (_) {}
          files.push(file);
        } else if (child.kind === 'directory') {
          const nested = await collectFilesFromDirectoryHandle(child, nextPath);
          files.push(...nested);
        }
      }
      return files;
    }

    async function pickFolderDeep() {
      if (!window.showDirectoryPicker) {
        document.getElementById('uploadResult').innerText = 'متصفحك لا يدعم اختيار المجلد العميق. استخدم السحب والإفلات.';
        return;
      }
      try {
        const root = await window.showDirectoryPicker();
        const allFiles = await collectFilesFromDirectoryHandle(root);
        appendFilesToSelection(allFiles);
        document.getElementById('uploadResult').innerText = `تمت إضافة ${allFiles.length} ملف من المجلد الرئيسي وكل المجلدات الفرعية.`;
      } catch (_) {
        // User canceled picker or permission denied.
      }
    }

    async function handleDrop(e) {
      e.preventDefault();
      const dropZone = document.getElementById('dropZone');
      if (dropZone) dropZone.classList.remove('active');

      const dt = e.dataTransfer;
      if (!dt) return;

      const filesToAdd = [];
      const items = Array.from(dt.items || []);
      if (items.length && items.some(i => typeof i.webkitGetAsEntry === 'function')) {
        for (const item of items) {
          const entry = item.webkitGetAsEntry ? item.webkitGetAsEntry() : null;
          if (!entry) continue;
          const files = await collectFilesFromEntry(entry);
          filesToAdd.push(...files);
        }
      } else {
        filesToAdd.push(...Array.from(dt.files || []));
      }

      appendFilesToSelection(filesToAdd);
      document.getElementById('uploadResult').innerText = 'تمت إضافة العناصر المسحوبة. اضغط رفع الملفات.';
    }

    function getHeaderIndexJs(headers, keywords) {
      const hs = (headers || []).map(h => normalizeTextForSearch(h));
      const ks = keywords.map(k => normalizeTextForSearch(k));
      for (let i = 0; i < hs.length; i++) {
        if (ks.includes(hs[i])) return i;
      }
      for (let i = 0; i < hs.length; i++) {
        if (ks.some(k => hs[i].includes(k))) return i;
      }
      return null;
    }

    function headerHasOriginalMarkerJs(h) {
      return h.includes('اصلي') || h.includes('أصلي') || h.includes('اصليه');
    }
    function headerBadForOriginalJs(h) {
      if (!h) return false;
      if (h.includes('متشابه') || h.includes('تشابه')) return !headerHasOriginalMarkerJs(h);
      if (h.includes('البدائل') || h.includes('بدائل')) return !headerHasOriginalMarkerJs(h);
      return false;
    }
    function headerBadForAltJs(h) {
      if (!h) return false;
      if (h.includes('رقم الشركات') || h.includes('رقم الشركة')) return true;
      if (h.includes('رقم شركة') && !h.includes('بديل') && !h.includes('متشابه') && !h.includes('بدائل') && !h.includes('تشابه')) return true;
      if (headerHasOriginalMarkerJs(h) && !h.includes('متشابه') && !h.includes('بديل') && !h.includes('بدائل') && !h.includes('تشابه')) return true;
      return false;
    }
    function headerBadForCompanyJs(h) {
      if (!h) return false;
      if (h.includes('رقم الشركات') || h.includes('رقم الشركة')) return false;
      if (h.includes('شركات') && h.includes('رقم')) return false;
      if (h.includes('شركة') && h.includes('رقم') && !h.includes('نوع')) return false;
      if (h.includes('نوع') && !h.includes('شركة') && !h.includes('شركات')) {
        if (h.includes('قطعه') || h.includes('قطعة') || h.includes('القطعه') || h.includes('القطعة') || h.includes('صنف')) return false;
        return true;
      }
      if (h === 'الصنف' || h === 'صنف' || (h.length <= 12 && h.includes('صنف') && !h.includes('رقم') && !h.includes('شرك'))) return true;
      return false;
    }

    function findHeaderColumnJs(headers, keywords, usedSet, minSubstringLen, rejectFn) {
      const hs = (headers || []).map(h => normalizeTextForSearch(h));
      const ks = [...new Set(keywords.map(k => normalizeTextForSearch(k)).filter(Boolean))]
        .sort((a, b) => b.length - a.length);
      const used = usedSet || new Set();
      for (let i = 0; i < hs.length; i++) {
        if (used.has(i) || !hs[i]) continue;
        if (rejectFn && rejectFn(hs[i])) continue;
        if (ks.includes(hs[i])) return i;
      }
      let bestIdx = null;
      let bestLen = -1;
      let bestI = 9999;
      for (let i = 0; i < hs.length; i++) {
        if (used.has(i) || !hs[i]) continue;
        const h = hs[i];
        if (rejectFn && rejectFn(h)) continue;
        for (const kw of ks) {
          if (kw.length < minSubstringLen) continue;
          if (h.includes(kw)) {
            if (kw.length > bestLen || (kw.length === bestLen && i < bestI)) {
              bestLen = kw.length;
              bestI = i;
              bestIdx = i;
            }
          }
        }
      }
      return bestIdx;
    }

    const HEADER_COMPANY_JS = ['رقم الشركات', 'رقم الشركة', 'رقم الشركه', 'رقم شركة', 'رقم شركات', 'ارقام الشركات', 'أرقام الشركات', 'ارقام شركات', 'أرقام شركات', 'ارقام الشركه', 'أرقام الشركه', 'company number', 'رقم الشركات الموحد', 'رقم المرجع', 'المرجع', 'مرجع الشركة', 'كود الشركة', 'كود الصنف', 'رقم الصنف التجاري', 'الصنف التجاري', 'الكود التجاري', 'كود المنتج', 'رقم المنتج التجاري', 'رقم التاجر', 'رقم المورد', 'كود التخزين', 'رقم التعريف', 'رقم مرجع', 'part number', 'reference', 'stock code', 'sku'];
    const HEADER_ORIGINAL_JS = ['الرقم الاصلي', 'الرقم الأصلي', 'رقم اصلي', 'رقم أصلي', 'ارقام اصلية', 'أرقام أصلية', 'ارقام أصلية', 'أرقام اصلية', 'original number', 'الرقم الاصلي للقطعة', 'الرقم الأصلي للقطعة'];
    const HEADER_ALT_JS = ['متشابهات', 'متشابهاب', 'المتشابهات', 'التشابهات', 'المشابهات', 'المشابهين', 'البدائل', 'بدائل', 'البديلة', 'السيارات البديلة', 'السيارة البديلة', 'سيارات مشابهة', 'سيارة مشابهة', 'alternatives', 'alternative', 'سيارات بديلة', 'السيارات المعادلة'];
    const HEADER_ITEM_NAME_JS = ['اسم السيارة', 'اسم الصنف', 'اسم السيارة/الصنف', 'النوع', 'نوع', 'نوع السيارة', 'نوع المركبة', 'نوع المركبه', 'الصنف', 'item name', 'name'];
    const HEADER_ITEM_NUMBER_JS = ['رقم الصنف', 'رقم القطعة', 'رقم الصنف الدولي', 'item number', 'part number'];
    const HEADER_NOTES_JS = ['ملاحظات', 'ملاحظة', 'notes'];

    function forwardFillHeaderRowJs(rawRow) {
      const cleaned = (rawRow || []).map(v => String(v || '').trim());
      const out = [];
      let carry = '';
      for (const s of cleaned) {
        if (s) {
          carry = s;
          out.push(s);
        } else {
          out.push(carry);
        }
      }
      return out;
    }

    function sanitizeAlternativesVsOriginalJs(alternatives, originalNumbers) {
      const a = String(alternatives || '').trim();
      if (!a) return '';
      const o = String(originalNumbers || '').trim();
      if (!o) return a;
      if (normalizeTextForSearch(a) === normalizeTextForSearch(o)) return '';
      if (normalizeTextCompactJs(a) === normalizeTextCompactJs(o)) return '';
      return a;
    }

    function normalizeHeaderCellForMatchJs(raw) {
      let t = normalizeTextForSearch(String(raw || ''));
      t = t.replace(/\u200c|\u200b|\ufeff|\xa0/g, '');
      return t.replace(/\s+/g, ' ').trim();
    }

    function findCompanyColumnByExactHeaderJs(header) {
      const canon = new Set([
        normalizeTextForSearch('رقم الشركات'),
        normalizeTextForSearch('رقم الشركة'),
        normalizeTextForSearch('رقم شركات'),
        normalizeTextForSearch('ارقام الشركات'),
        normalizeTextForSearch('أرقام الشركات'),
        normalizeTextForSearch('ارقام الشركة'),
        normalizeTextForSearch('أرقام الشركة')
      ]);
      for (let i = 0; i < header.length; i++) {
        const h = normalizeHeaderCellForMatchJs(header[i]);
        if (!h) continue;
        if (canon.has(h)) return i;
        if (h.length <= 32 && h.includes('رقم') && (h.includes('شركات') || h.endsWith('شركه')) && !headerBadForCompanyJs(h)) return i;
      }
      return null;
    }

    function resolveProductColumnIndicesJs(header) {
      const used = new Set();
      function assign(kws, minSub, rejectFn) {
        const idx = findHeaderColumnJs(header, kws, used, minSub, rejectFn);
        if (idx !== null && idx !== undefined) used.add(idx);
        return idx;
      }
      let companyIdx = findCompanyColumnByExactHeaderJs(header);
      if (companyIdx !== null && companyIdx !== undefined) {
        used.add(companyIdx);
      } else {
        companyIdx = assign(HEADER_COMPANY_JS, 6, headerBadForCompanyJs);
      }
      const originalIdx = assign(HEADER_ORIGINAL_JS, 6, headerBadForOriginalJs);
      let altIdx = assign(HEADER_ALT_JS, 5, headerBadForAltJs);
      const itemNameIdx = assign(HEADER_ITEM_NAME_JS, 4);
      const notesIdx = assign(HEADER_NOTES_JS, 4);
      let itemNumIdx = findHeaderColumnJs(header, HEADER_ITEM_NUMBER_JS, used, 6);
      if (itemNumIdx !== null && itemNumIdx !== undefined) used.add(itemNumIdx);
      else if (companyIdx !== null && companyIdx !== undefined) itemNumIdx = companyIdx;
      else itemNumIdx = 4;
      if (originalIdx != null && originalIdx !== undefined && altIdx === originalIdx) altIdx = null;
      if (companyIdx === null || companyIdx === undefined) {
        const skipLoose = new Set([originalIdx, altIdx, notesIdx].filter((x) => x !== null && x !== undefined));
        for (let i = 0; i < header.length; i++) {
          if (skipLoose.has(i)) continue;
          const h = normalizeTextForSearch(String(header[i] || ''));
          if (!h) continue;
          if ((h.includes('رقم') || h.includes('كود') || h.includes('ارقام')) && (h.includes('شركات') || h.includes('شركه'))) {
            if (headerBadForCompanyJs(h)) continue;
            companyIdx = i;
            break;
          }
        }
      }
      if ((companyIdx === null || companyIdx === undefined) && originalIdx != null && originalIdx !== undefined) {
        const blocked = new Set([itemNameIdx, altIdx, originalIdx, notesIdx].filter((x) => x !== null && x !== undefined));
        for (const delta of [1, -1, 2, -2]) {
          const fb = originalIdx + delta;
          if (fb < 0 || blocked.has(fb) || fb >= header.length) continue;
          companyIdx = fb;
          break;
        }
      }
      return { companyIdx, originalIdx, altIdx, itemNameIdx, itemNumIdx, notesIdx };
    }

    function pickItemNameJs(row, preferredIdx) {
      const val = (i) => (i !== null && i !== undefined && i >= 0 ? String(row[i] || '').trim() : '');
      const primary = val(preferredIdx);
      if (primary) return primary;

      for (const i of [1, 0, 2, 3, 4, 5]) {
        const v = val(i);
        if (!v) continue;
        if (/[A-Za-z\u0600-\u06FF]/.test(v)) return v;
      }
      for (let i = 0; i < Math.min(10, row.length || 0); i++) {
        const v = val(i);
        if (v) return v;
      }
      return '';
    }

    async function sha256Hex(arrayBuffer) {
      const hash = await crypto.subtle.digest('SHA-256', arrayBuffer);
      const bytes = Array.from(new Uint8Array(hash));
      return bytes.map(b => b.toString(16).padStart(2, '0')).join('');
    }

    async function readJsonSafe(res) {
      const raw = await res.text();
      let data = {};
      try {
        data = raw ? JSON.parse(raw) : {};
      } catch (_) {
        data = { detail: raw || `HTTP ${res.status}` };
      }
      return data;
    }

    async function parseExcelRowsForChunkUpload(file) {
      const arrayBuffer = await file.arrayBuffer();
      const wb = XLSX.read(arrayBuffer, { type: 'array' });
      const sourceKey = sourceKeyForFile(file);
      const rows = [];
      const itemNamesSet = new Set();

      for (const sheetName of wb.SheetNames) {
        const ws = wb.Sheets[sheetName];
        const grid = XLSX.utils.sheet_to_json(ws, { header: 1, defval: '' });
        const headerCandidates = ['اسم الصنف', 'اسم السيارة', 'رقم الصنف', 'رقم الشركة', 'رقم الشركات', 'الرقم الاصلي', 'البدائل', 'متشابهات'];
        function previewRowIsLikelyHeaderJs(cells) {
          const nonempty = cells.map(s => String(s || '').trim()).filter(Boolean);
          if (nonempty.length < 2) return true;
          let dataLike = 0;
          for (const v of nonempty.slice(0, 14)) {
            if (v.length < 5) continue;
            const dig = [...v].filter(ch => /\d/.test(ch)).length;
            if (v.length >= 10 && dig / v.length > 0.4) dataLike++;
            else if (v.length >= 6 && dig / v.length > 0.55) dataLike++;
          }
          return dataLike < Math.max(2, Math.floor(nonempty.length / 3));
        }
        let headerRowIndex = 0;
        let bestScore = -10;
        for (let i = 0; i < Math.min(30, grid.length); i++) {
          const rowCells = (grid[i] || []).map(v => String(v || '').trim());
          let score = 0;
          for (const k of headerCandidates) {
            if (getHeaderIndexJs(rowCells, [k]) !== null) score += 1;
          }
          if (!previewRowIsLikelyHeaderJs(rowCells)) score -= 12;
          if (score > bestScore) {
            bestScore = score;
            headerRowIndex = i;
          }
        }
        const header = forwardFillHeaderRowJs(grid[headerRowIndex] || []);
        const colMap = resolveProductColumnIndicesJs(header);
        const itemNameIdx = colMap.itemNameIdx != null ? colMap.itemNameIdx : 0;
        const itemNumberIdx = colMap.itemNumIdx != null ? colMap.itemNumIdx : 4;
        const companyIdx = colMap.companyIdx;
        const originalIdx = colMap.originalIdx;
        const notesIdx = colMap.notesIdx;
        const altIdx = colMap.altIdx != null && colMap.altIdx !== undefined ? colMap.altIdx : null;

        for (const r of grid.slice(headerRowIndex + 1)) {
          const itemName = pickItemNameJs(r, itemNameIdx);
          const itemNumber = String(r[itemNumberIdx] || '').trim();
          let companyNumber = String((companyIdx != null && companyIdx >= 0) ? r[companyIdx] || '' : '').trim();
          const originalNumbers = String((originalIdx != null && originalIdx >= 0) ? r[originalIdx] || '' : '').trim();
          if (!companyNumber && originalNumbers && originalIdx != null && originalIdx !== undefined) {
            const oC = normalizeTextCompactJs(originalNumbers);
            const skipAdj = new Set([companyIdx, altIdx, notesIdx, itemNameIdx, itemNumberIdx].filter((x) => x != null && x !== undefined));
            for (const delta of [1, -1]) {
              const j = originalIdx + delta;
              if (skipAdj.has(j) || j < 0) continue;
              const cand = String(r[j] || '').trim();
              if (!cand || normalizeTextCompactJs(cand) === oC) continue;
              companyNumber = cand;
              break;
            }
          }
          const notes = String((notesIdx != null && notesIdx >= 0) ? r[notesIdx] || '' : '').trim();
          let alternatives = String((altIdx != null && altIdx >= 0) ? r[altIdx] || '' : '').trim();
          alternatives = sanitizeAlternativesVsOriginalJs(alternatives, originalNumbers);
          const sizeWidth = String(r[1] || '').trim();
          const sizeDiameter = String(r[2] || '').trim();
          const sizeHeight = String(r[3] || '').trim();

          if (!itemName && !itemNumber && !companyNumber && !originalNumbers && !notes && !alternatives && !sizeWidth && !sizeDiameter && !sizeHeight) continue;

          rows.push({
            item_name: itemName,
            item_number: itemNumber,
            company_number: companyNumber,
            original_numbers: originalNumbers,
            notes: notes,
            alternatives: alternatives,
            size_width: sizeWidth,
            size_diameter: sizeDiameter,
            size_height: sizeHeight,
            source_sheet: String(sheetName)
          });
          if (itemName) itemNamesSet.add(itemName);
        }
      }

      return {
        file_name: sourceKey,
        file_size: arrayBuffer.byteLength,
        content_hash: await sha256Hex(arrayBuffer),
        rows,
        item_names_count: itemNamesSet.size
      };
    }

    async function uploadLargeFileChunked(file) {
      const parsed = await parseExcelRowsForChunkUpload(file);
      try {
        const startRes = await fetch('/upload_rows/start', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({
            file_name: parsed.file_name,
            content_hash: parsed.content_hash,
            file_size: parsed.file_size
          })
        });
        const startData = await readJsonSafe(startRes);
        if (!startRes.ok) throw new Error(startData.detail || 'start-failed');
        if (startData.skip) return { files_count: 0, skipped_files: 1, inserted_rows: 0, inserted_items: 0 };

        const chunkSize = 1200;
        let inserted = 0;
        for (let i = 0; i < parsed.rows.length; i += chunkSize) {
          const chunk = parsed.rows.slice(i, i + chunkSize);
          const chunkRes = await fetch('/upload_rows/chunk', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ file_name: parsed.file_name, rows: chunk })
          });
          const chunkData = await readJsonSafe(chunkRes);
          if (!chunkRes.ok) throw new Error(chunkData.detail || 'chunk-failed');
          inserted += Number(chunkData.inserted_rows || 0);
        }

        const finishRes = await fetch('/upload_rows/finish', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({
            file_name: parsed.file_name,
            content_hash: parsed.content_hash,
            file_size: parsed.file_size
          })
        });
        const finishData = await readJsonSafe(finishRes);
        if (!finishRes.ok) throw new Error(finishData.detail || 'finish-failed');

        return {
          files_count: 1,
          skipped_files: 0,
          inserted_rows: Number(finishData.rows_count || inserted),
          inserted_items: Number(parsed.item_names_count || 0)
        };
      } catch (err) {
        // Prevent partial file data if a chunk fails mid-upload.
        try {
          await fetch('/upload_rows/abort', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ file_name: parsed.file_name })
          });
        } catch (_) {}
        throw err;
      }
    }

    function normalizeTextForSearch(text) {
      return String(text || '')
        .replace(/[٠-٩]/g, d => String('٠١٢٣٤٥٦٧٨٩'.indexOf(d)))
        .replace(/[۰-۹]/g, d => String('۰۱۲۳۴۵۶۷۸۹'.indexOf(d)))
        .replace(/[أإآٱ]/g, 'ا')
        .replace(/ى/g, 'ي')
        .replace(/ؤ/g, 'و')
        .replace(/ئ/g, 'ي')
        .replace(/ة/g, 'ه')
        .replace(/ء/g, '')
        .replace(/ـ/g, '')
        .replace(/[\u064B-\u065F\u0670\u06D6-\u06ED]/g, '')
        .trim()
        .toLowerCase()
        .replace(/[\\s\\-_]+/g, ' ');
    }

    function normalizeTextCompactJs(text) {
      return normalizeTextForSearch(text).replace(/[\\s\\-_/]+/g, '');
    }

    function tokenizeQueryForSearch(text) {
      let raw = normalizeTextForSearch(text).replace(/[،,;/|]+/g, ' ');
      raw = raw.replace(/[()]+/g, ' ');
      raw = raw.replace(/(\d{1,2})\s*[.,،]\s*(\d{1,3})(?!\d)/g, '$1.$2');
      raw = raw.replace(/\s+/g, ' ').trim();
      return raw.split(' ').filter(Boolean).filter(t => !['+', '-', '–', '—'].includes(t));
    }

    function splitQueryDisplayWordsJs(text) {
      let raw = String(text || '')
        .replace(/[٠-٩]/g, d => String('٠١٢٣٤٥٦٧٨٩'.indexOf(d)))
        .replace(/[۰-۹]/g, d => String('۰۱۲۳۴۵۶۷۸۹'.indexOf(d)))
        .trim();
      raw = raw.replace(/[،,;/|]+/g, ' ');
      raw = raw.replace(/[()]+/g, ' ');
      raw = raw.replace(/(\d{1,2})\s*[.,،]\s*(\d{1,3})(?!\d)/g, '$1.$2');
      raw = raw.replace(/\s+/g, ' ').trim();
      return raw.split(' ').filter(Boolean).filter(t => !['+', '-', '–', '—'].includes(t));
    }

    function displayFormsForTextTokensJs(textTokens, rawQuery) {
      const words = splitQueryDisplayWordsJs(rawQuery);
      const used = new Set();
      return textTokens.map(t => {
        const nt = normalizeTextForSearch(t);
        for (let i = 0; i < words.length; i++) {
          if (used.has(i)) continue;
          if (normalizeTextForSearch(words[i]) === nt) {
            used.add(i);
            return words[i];
          }
        }
        return t;
      });
    }

    function formatSizeLabelDisplayJs(sourceFile, width, diameter, height) {
      const sw = String(width || '').trim();
      const sd = String(diameter || '').trim();
      const sh = String(height || '').trim();
      if (!isSizeFileNameJs(sourceFile)) return '';
      const sf = normalizeTextForSearch(sourceFile || '');
      if (normalizedPathMatchesDriveshaftJs(sf) || (isSizeFileNameJs(sourceFile) && !sh)) {
        const p = [sw, sd].filter(Boolean);
        return p.length ? p.join(' · ') : '';
      }
      const p2 = [sw, sd, sh].filter(Boolean);
      return p2.length ? p2.join(' · ') : '';
    }

    function normalizedPathMatchesDriveshaftJs(sf) {
      if (!sf) return false;
      const t = normalizeTextForSearch;
      if (sf.includes(t('مصلبة دراي شفط')) || sf.includes(t('مصلبات دراي شفط')) || sf.includes(t('مصلب دراي شفط'))) return true;
      if (!sf.includes('مصلب')) return false;
      const hasShaft = sf.includes('شفط') || sf.includes('shaft');
      const hasDrive = sf.includes('دراي') || sf.includes('دري') || sf.includes('dry') || sf.includes('drive');
      return hasShaft || hasDrive;
    }

    function isDriveshaftSizeTypeJs(st) {
      if (!st) return false;
      const t = normalizeTextForSearch;
      const keys = [t('مصلبات دراي شفط'), t('مصلبه دراي شفط'), t('مصلبات'), t('مصلب دراي شفط'), t('مصلب دراي')];
      if (keys.includes(st)) return true;
      return st.includes('مصلب') && (st.includes('شفط') || st.includes('دراي') || st.includes('دري') || st.includes('dry') || st.includes('drive') || st.includes('shaft'));
    }

    function isSizeFileNameJs(sourceFile) {
      const sf = normalizeTextForSearch(sourceFile || '');
      return sf.includes('لبادات')
        || sf.includes('بيل')
        || sf.includes('بيلية')
        || sf.includes('نابات')
        || normalizedPathMatchesDriveshaftJs(sf);
    }

    function fileMatchesSizeTypeJs(sourceFile, sizeType) {
      const st = normalizeTextForSearch(sizeType || '');
      if (!st) return true;
      const sf = normalizeTextForSearch(sourceFile || '');
      if (st === 'لبادات') return sf.includes('لبادات');
      if (['بيلية', 'بيل', 'بلي'].includes(st)) {
        return sf.includes(normalizeTextForSearch('نابات + بيل عجل'))
          || sf.includes('بيل مشكلة')
          || sf.includes(normalizeTextForSearch('بيـل مشكلة'));
      }
      if (isDriveshaftSizeTypeJs(st)) return normalizedPathMatchesDriveshaftJs(sf);
      return true;
    }

    function fileMatchesSizeTypeWithFallbackJs(r, sizeType, sw, sd) {
      const st = normalizeTextForSearch(sizeType || '');
      if (!st) return true;
      if (fileMatchesSizeTypeJs(r.source_file || '', sizeType)) return true;
      if (!isDriveshaftSizeTypeJs(st)) return false;
      const sf = normalizeTextForSearch(r.source_file || '');
      if (sf.includes('لبادات')) return false;
      if (sf.includes('بيل مشكلة') || sf.includes(normalizeTextForSearch('بيـل مشكلة')) || sf.includes(normalizeTextForSearch('نابات + بيل عجل'))) return false;
      const swq = normalizeTextForSearch(sw || '').trim();
      const sdq = normalizeTextForSearch(sd || '').trim();
      if (!swq || !sdq) return false;
      const rw = String(r.size_width || '').trim();
      const rd = String(r.size_diameter || '').trim();
      if (!rw || !rd) return false;
      return (cellMatchesSizeQueryDriveshaftJs(rw, sw) && cellMatchesSizeQueryDriveshaftJs(rd, sd))
        || (cellMatchesSizeQueryDriveshaftJs(rw, sd) && cellMatchesSizeQueryDriveshaftJs(rd, sw));
    }

    function rowMatchesSizeFiltersJs(r, sizeType, sw, sd, sh) {
      const st = normalizeTextForSearch(sizeType || '');
      const wv = r.size_width || '';
      const dv = r.size_diameter || '';
      const hv = r.size_height || '';
      const swq = normalizeTextForSearch(sw || '').trim();
      const sdq = normalizeTextForSearch(sd || '').trim();
      if (isDriveshaftSizeTypeJs(st)) {
        let dimOk;
        if (swq && sdq) {
          dimOk = (cellMatchesSizeQueryDriveshaftJs(wv, sw) && cellMatchesSizeQueryDriveshaftJs(dv, sd))
            || (cellMatchesSizeQueryDriveshaftJs(wv, sd) && cellMatchesSizeQueryDriveshaftJs(dv, sw));
        } else if (swq) {
          dimOk = cellMatchesSizeQueryDriveshaftJs(wv, sw) || cellMatchesSizeQueryDriveshaftJs(dv, sw);
        } else if (sdq) {
          dimOk = cellMatchesSizeQueryDriveshaftJs(wv, sd) || cellMatchesSizeQueryDriveshaftJs(dv, sd);
        } else {
          dimOk = true;
        }
        return dimOk && cellMatchesSizeQueryJs(hv, sh);
      }
      return cellMatchesSizeQueryJs(wv, sw) && cellMatchesSizeQueryJs(dv, sd) && cellMatchesSizeQueryJs(hv, sh);
    }

    function cellMatchesSizeQueryJs(cellVal, queryVal) {
      if (!String(queryVal || '').trim()) return true;
      const q = normalizeTextForSearch(queryVal).replace(',', '.').trim();
      const c = normalizeTextForSearch(cellVal).replace(',', '.').trim();
      if (c === q) return true;
      const nq = Number(q);
      const nc = Number(c);
      if (!Number.isNaN(nq) && !Number.isNaN(nc) && nq === nc) return true;
      if (Number.isNaN(nq)) return c.includes(q);
      const nums = [];
      const re = /\\d+(?:\\.\\d+)?/g;
      let m;
      while ((m = re.exec(c)) !== null) {
        const v = Number(m[0]);
        if (!Number.isNaN(v)) nums.push(v);
      }
      return nums.some(v => v === nq);
    }

    function cellMatchesSizeQueryDriveshaftJs(cellVal, queryVal) {
      if (cellMatchesSizeQueryJs(cellVal, queryVal)) return true;
      const q = normalizeTextForSearch(queryVal || '').replace(',', '.').trim();
      if (!/^\\d$/.test(q)) return false;
      const t = normalizeTextForSearch(cellVal || '').replace(',', '.');
      const re = /\\d+(?:\\.\\d+)?/g;
      let m;
      while ((m = re.exec(t)) !== null) {
        const s = m[0];
        const sn = s.replace(/^0+/, '') || '0';
        if (s.startsWith(q) || sn.startsWith(q)) return true;
      }
      return false;
    }

    function updateSizeInputsLayout() {
      const type = (document.getElementById('sizeTypeInput').value || '').trim();
      const width = document.getElementById('sizeWidthInput');
      const diameter = document.getElementById('sizeDiameterInput');
      const height = document.getElementById('sizeHeightInput');
      if (isDriveshaftSizeTypeJs(normalizeTextForSearch(type))) {
        width.placeholder = 'الراسية (B)';
        diameter.placeholder = 'طول المصلبة (C)';
        height.style.display = 'none';
        height.value = '';
      } else {
        width.placeholder = 'العرض (B)';
        diameter.placeholder = 'القطر (C)';
        height.style.display = '';
        height.placeholder = 'الارتفاع (D)';
      }
    }

    function extractFileSearchHintsJs(query, tokens) {
      const qn = normalizeTextForSearch(query);
      const hints = [];
      const consumed = new Set();
      const consume = (...parts) => parts.forEach(p => consumed.add(normalizeTextForSearch(p)));

      if ((qn.includes('فلتر') || qn.includes('فلاتر')) && qn.includes('هواء')) {
        hints.push('فلتر هواء', 'فلاتر هواء');
        consume('فلتر', 'فلاتر', 'هواء');
      }
      if ((qn.includes('فلتر') || qn.includes('فلاتر')) && qn.includes('زيت')) {
        hints.push('فلتر زيت', 'فلاتر زيت');
        consume('فلتر', 'فلاتر', 'زيت');
      }
      if ((qn.includes('فلتر') || qn.includes('فلاتر')) && (qn.includes('سولار') || qn.includes('ديزل'))) {
        hints.push('فلتر سولار', 'فلاتر السولار', 'فلاتر سولار', 'فلتر ديزل', 'فلاتر ديزل');
        consume('فلتر', 'فلاتر', 'سولار', 'ديزل', 'السولار');
      }

      const cleanedTokens = (tokens || []).filter(t => !consumed.has(normalizeTextForSearch(t)));
      const uniqueHints = Array.from(new Set(hints.map(normalizeTextForSearch))).filter(Boolean);
      return { fileHints: uniqueHints, textTokens: cleanedTokens };
    }

    function fileSearchHintMatchesPathJs(pathNorm, hintNorm) {
      const pn = String(pathNorm || '').trim();
      const hn = String(hintNorm || '').trim();
      if (!hn) return true;
      if (pn.includes(hn)) return true;
      const significant = hn.split(/\s+/).filter(w => w.length >= 2);
      if (significant.length < 2) return pn.includes(hn);
      function wordOk(w) {
        if (pn.includes(w)) return true;
        if (w.startsWith('ال') && w.length > 2 && pn.includes(w.slice(2))) return true;
        return false;
      }
      return significant.every(wordOk);
    }

    function parseYearTokenJs(token) {
      const cleaned = String(token || '').replace(/[^\\d]/g, '');
      if (!(cleaned.length === 2 || cleaned.length === 4)) return null;
      const year = Number(cleaned);
      if (cleaned.length === 2) return year <= 30 ? 2000 + year : 1900 + year;
      return year;
    }

    function parseQueryYearTokenJs(token) {
      const raw = normalizeTextForSearch(token);
      const currentYear = new Date().getFullYear();

      if (/^\d{1,2}\.\d{1,3}$/.test(raw)) return null;

      if (/^(?:\+\d{2,4}|\d{2,4}\+)$/.test(raw)) {
        const start = parseYearTokenJs(raw.startsWith('+') ? raw.slice(1) : raw.slice(0, -1));
        if (!start) return null;
        return Array.from({ length: currentYear - start + 1 }, (_, i) => start + i);
      }
      const rangeMatch = raw.match(/^(\d{2,4})\s*[-–—]\s*(\d{2,4})$/);
      if (rangeMatch) {
        const start = parseYearTokenJs(rangeMatch[1]);
        const end = parseYearTokenJs(rangeMatch[2]);
        if (!start || !end) return null;
        const low = Math.min(start, end);
        const high = Math.max(start, end);
        return Array.from({ length: high - low + 1 }, (_, i) => low + i);
      }
      if (/^(?:-\d{2,4}|\d{2,4}-)$/.test(raw)) {
        const end = parseYearTokenJs(raw.startsWith('-') ? raw.slice(1) : raw.slice(0, -1));
        if (!end) return null;
        return Array.from({ length: end - 1900 + 1 }, (_, i) => 1900 + i);
      }
      if (/^\d{2,4}$/.test(raw)) {
        const y = parseYearTokenJs(raw);
        return y ? [y] : null;
      }
      return null;
    }

    function isExplicitYearOperatorTokenJs(token) {
      const raw = normalizeTextForSearch(token);
      return /^(?:\+\d{2,4}|\d{2,4}\+)$/.test(raw)
        || /^(\d{2,4})\s*[-–—]\s*(\d{2,4})$/.test(raw)
        || /^(?:-\d{2,4}|\d{2,4}-)$/.test(raw);
    }

    function yearInRangeTextJs(year, text) {
      const normalized = normalizeTextForSearch(text);
      if (!normalized) return false;

      const currentYear = new Date().getFullYear();
      for (const m of normalized.matchAll(/(?:^|[^\d])(\d{4})(?!\d)/g)) {
        if (Number(m[1]) === year) return true;
      }
      for (const m of normalized.matchAll(/(?:^|[^\d])(\d{2})(?!\d)/g)) {
        if (parseYearTokenJs(m[1]) === year) return true;
      }

      // Chained plus markers: +03,+08 => 03-07 and +08.
      const plusMarks = [];
      for (const m of normalized.matchAll(/(?:\+\s*(\d{2,4})(?!\d)|(?:^|[^\d])(\d{2,4})\s*\+(?!\d))/g)) {
        const token = m[1] || m[2];
        const start = parseYearTokenJs(token);
        if (start) plusMarks.push({ idx: m.index ?? 0, start });
      }
      if (plusMarks.length) {
        plusMarks.sort((a, b) => a.idx - b.idx);
        for (let i = 0; i < plusMarks.length; i++) {
          const start = plusMarks[i].start;
          const end = i + 1 < plusMarks.length ? plusMarks[i + 1].start - 1 : currentYear;
          if (start <= year && year <= end) return true;
        }
      }

      for (const m of normalized.matchAll(/\+\s*(\d{2,4})(?!\d)/g)) {
        const start = parseYearTokenJs(m[1]);
        if (start && start <= year && year <= currentYear) return true;
      }
      for (const m of normalized.matchAll(/(?:^|[^\d])(\d{2,4})\s*\+(?!\d)/g)) {
        const start = parseYearTokenJs(m[1]);
        if (start && start <= year && year <= currentYear) return true;
      }
      for (const m of normalized.matchAll(/(?:^|[^\d])(\d{2,4})\s*[-–—]\s*(\d{2,4})(?!\d)/g)) {
        const start = parseYearTokenJs(m[1]);
        const end = parseYearTokenJs(m[2]);
        if (!start || !end) continue;
        const low = Math.min(start, end);
        const high = Math.max(start, end);
        if (low <= year && year <= high) return true;
      }

      // Chained minus markers: -03,-08 => <=03 and 04-08.
      const minusMarks = [];
      for (const m of normalized.matchAll(/(?:[-–—]\s*(\d{2,4})(?!\d)|(?:^|[^\d])(\d{2,4})\s*[-–—](?!\s*\d))/g)) {
        const token = m[1] || m[2];
        const end = parseYearTokenJs(token);
        if (end) minusMarks.push({ idx: m.index ?? 0, end });
      }
      if (minusMarks.length) {
        minusMarks.sort((a, b) => a.idx - b.idx);
        if (year <= minusMarks[0].end) return true;
        for (let i = 1; i < minusMarks.length; i++) {
          const low = minusMarks[i - 1].end + 1;
          const high = minusMarks[i].end;
          if (low <= high && low <= year && year <= high) return true;
        }
      }

      for (const m of normalized.matchAll(/[-–—]\s*(\d{2,4})(?!\d)/g)) {
        const end = parseYearTokenJs(m[1]);
        if (end && year <= end) return true;
      }
      for (const m of normalized.matchAll(/(?:^|[^\d])(\d{2,4})\s*[-–—](?!\s*\d)/g)) {
        const end = parseYearTokenJs(m[1]);
        if (end && year <= end) return true;
      }
      return false;
    }

    function splitAlternativeSegmentsJs(alternatives) {
      return String(alternatives || '').split(/[\\/|,\\n;]+/).map(s => s.trim()).filter(Boolean);
    }

    function normSegmentWordSetJs(nseg) {
      const edges = '.,|;:()[]';
      const out = new Set();
      for (const w of String(nseg || '').split(/\s+/)) {
        let t = w;
        while (t.length && edges.includes(t[0])) t = t.slice(1);
        while (t.length && edges.includes(t[t.length - 1])) t = t.slice(0, -1);
        if (t) out.add(t);
      }
      return out;
    }

    function tokenInNormSegmentJs(ntok, nseg) {
      if (!ntok) return false;
      const wset = normSegmentWordSetJs(nseg);
      if (ntok.length <= 2) return wset.has(ntok);
      if (wset.has(ntok)) return true;
      return nseg.includes(ntok);
    }

    function tokenAppearsInAlternativesColumnJs(ntok, alternatives) {
      const nAlt = normalizeTextForSearch(alternatives);
      if (!ntok) return false;
      const wset = normSegmentWordSetJs(nAlt);
      if (ntok.length <= 2) return wset.has(ntok);
      if (wset.has(ntok)) return true;
      return nAlt.includes(ntok);
    }

    function tokensMatchingAlternativesColumnJs(textTokens, alternatives) {
      const seen = new Set();
      const out = [];
      for (const t of textTokens) {
        const nt = normalizeTextForSearch(t);
        if (!nt || seen.has(nt)) continue;
        if (!tokenAppearsInAlternativesColumnJs(nt, alternatives)) continue;
        seen.add(nt);
        out.push(nt);
      }
      return out;
    }

    function isWordCharAltJs(ch) {
      if (!ch) return false;
      const c = ch[0];
      if (c >= '0' && c <= '9') return true;
      if ((c >= 'a' && c <= 'z') || (c >= 'A' && c <= 'Z')) return true;
      if (c >= '\u0600' && c <= '\u06FF') return true;
      return c === '+' || c === '.';
    }

    function findWholeWordSpanJs(haystack, needle) {
      if (!needle || !haystack) return null;
      const h = haystack.toLowerCase();
      const n = needle.toLowerCase();
      let pos = 0;
      while (true) {
        const idx = h.indexOf(n, pos);
        if (idx < 0) return null;
        const end = idx + n.length;
        const leftOk = idx === 0 || !isWordCharAltJs(haystack[idx - 1]);
        const rightOk = end >= haystack.length || !isWordCharAltJs(haystack[end]);
        if (leftOk && rightOk) return [idx, end];
        pos = idx + 1;
      }
    }

    function earliestTokenMatchSpanJs(alternatives, tokens) {
      let best = null;
      for (const tok of tokens) {
        const sp = findWholeWordSpanJs(alternatives, tok);
        if (!sp) continue;
        if (!best || sp[0] < best[0]) best = sp;
      }
      return best;
    }

    function sliceAlternativesFirstMatchToSlashJs(alternatives, tokens) {
      const sp = earliestTokenMatchSpanJs(alternatives, tokens);
      if (!sp) return null;
      const st = sp[0];
      const slashAt = alternatives.indexOf('/', st);
      if (slashAt < 0) return alternatives.slice(st).trim();
      return alternatives.slice(st, slashAt).trim();
    }

    function extractMatchedAlternativeJs(alternatives, textTokens, years, rawQuery = '') {
      void rawQuery;
      if (!String(alternatives || '').trim()) return '';
      const segments = splitAlternativeSegmentsJs(alternatives);
      if (!segments.length) return '';

      const altToks = tokensMatchingAlternativesColumnJs(textTokens, alternatives);
      let tokensUse;
      if (altToks.length) tokensUse = altToks;
      else if (years.length) tokensUse = [];
      else tokensUse = textTokens.map(normalizeTextForSearch).filter(Boolean);

      if (tokensUse.length) {
        const sliced = sliceAlternativesFirstMatchToSlashJs(String(alternatives), tokensUse);
        if (sliced) {
          if (!years.length || years.some(y => yearInRangeTextJs(y, sliced))) {
            return sliced;
          }
        }
      }

      function segMatchesTokens(n) {
        if (!tokensUse.length) return true;
        return tokensUse.every(t => tokenInNormSegmentJs(t, n));
      }

      const candidates = [];
      for (const segment of segments) {
        const n = normalizeTextForSearch(segment);
        if (!segMatchesTokens(n)) continue;
        const hasYear = !years.length || years.some(y => yearInRangeTextJs(y, segment));
        if (hasYear) candidates.push(segment);
      }
      if (candidates.length) {
        return candidates.reduce((a, b) => (a.length >= b.length ? a : b)).trim();
      }

      let pool = segments.slice();
      if (years.length) {
        const yp = segments.filter(s => years.some(y => yearInRangeTextJs(y, s)));
        if (yp.length) pool = yp;
      }

      if (tokensUse.length) {
        let bestSeg = '';
        let bestScore = -1;
        for (const segment of pool) {
          const n = normalizeTextForSearch(segment);
          const sc = tokensUse.filter(t => tokenInNormSegmentJs(t, n)).length;
          if (sc > bestScore || (sc === bestScore && segment.length > bestSeg.length)) {
            bestScore = sc;
            bestSeg = segment;
          }
        }
        if (bestScore > 0) return bestSeg.trim();
      }

      if (years.length) {
        const yearHits = segments.filter(s => years.some(y => yearInRangeTextJs(y, s)));
        if (yearHits.length) {
          return yearHits.reduce((a, b) => (a.length >= b.length ? a : b)).trim();
        }
        return '';
      }
      return '';
    }

    function yearMatchScoreJs(segment, years) {
      if (!segment || !years.length) return 0;
      const n = normalizeTextForSearch(segment);
      let best = 0;
      for (const y of years) {
        if ((new RegExp(`(^|\\\\D)${y}(\\\\D|$)`)).test(n)) best = Math.max(best, 6000);
      }
      for (const m of n.matchAll(/\+\s*(\d{2,4})(?!\d)/g)) {
        const start = parseYearTokenJs(m[1]);
        if (start && years.every(y => start <= y)) best = Math.max(best, 4000 + start);
      }
      for (const m of n.matchAll(/(?:^|[^\d])(\d{2,4})\s*\+(?!\d)/g)) {
        const start = parseYearTokenJs(m[1]);
        if (start && years.every(y => start <= y)) best = Math.max(best, 4000 + start);
      }
      for (const m of n.matchAll(/(?:^|[^\d])(\d{2,4})\s*[-–—]\s*(\d{2,4})(?!\d)/g)) {
        const a = parseYearTokenJs(m[1]); const b = parseYearTokenJs(m[2]);
        if (!a || !b) continue;
        const low = Math.min(a, b), high = Math.max(a, b);
        if (years.every(y => low <= y && y <= high)) best = Math.max(best, 5000 - (high - low));
      }
      return best;
    }

    function rowMatchesQueryJs(row, textTokens, years) {
      // كل توكن يجب أن يظهر في الصف (اسم/بدائل/ملف) — ترتيب الكلمات في الاستعلام لا يؤثر.
      const itemName = row.item_name || '';
      const alternatives = row.alternatives || '';
      const sourceFile = row.source_file || '';
      const itemNorm = normalizeTextForSearch(itemName);
      const altNorm = normalizeTextForSearch(alternatives);
      const fileNorm = normalizeTextForSearch(sourceFile);
      const combined = `${itemNorm} ${altNorm} ${fileNorm}`;
      const combinedCompact = normalizeTextCompactJs(`${itemName} ${alternatives} ${sourceFile}`);
      for (const token of textTokens) {
        const nt = normalizeTextForSearch(token);
        const nct = normalizeTextCompactJs(token);
        if (!combined.includes(nt) && !(nct && combinedCompact.includes(nct))) return false;
      }
      if (years.length) {
        const itemTokens = textTokens.map(normalizeTextForSearch).filter(t => itemNorm.includes(t));
        const altTokens = textTokens.map(normalizeTextForSearch).filter(t => altNorm.includes(t));
        const fileTokens = textTokens.map(normalizeTextForSearch).filter(t => fileNorm.includes(t));
        const itemYearOk = itemTokens.length > 0 && years.some(y => yearInRangeTextJs(y, itemName));
        let altYearOk = false;
        for (const seg of splitAlternativeSegmentsJs(alternatives)) {
          const n = normalizeTextForSearch(seg);
          const hasAltTokens = !altTokens.length || altTokens.every(t => n.includes(t));
          const hasYear = years.some(y => yearInRangeTextJs(y, seg));
          if (hasAltTokens && hasYear) { altYearOk = true; break; }
        }
        if (altTokens.length ? !altYearOk : !(itemYearOk || altYearOk)) return false;
        if (!altTokens.length && fileTokens.length && !itemTokens.length) {
          // File-context token match: don't require year context in row text.
        } else if (altTokens.length ? !altYearOk : !(itemYearOk || altYearOk)) {
          return false;
        }
      }
      return true;
    }

    function rowSearchRelevanceScoreJs(row, textTokens, years) {
      const item = String(row.item_name || '');
      const alt = String(row.alternatives || '');
      const bag = normalizeTextForSearch(`${item} ${alt}`);
      const bagC = normalizeTextCompactJs(`${item} ${alt}`);
      const altN = normalizeTextForSearch(alt);
      const itemN = normalizeTextForSearch(item);
      const altC = normalizeTextCompactJs(alt);
      const itemC = normalizeTextCompactJs(item);
      let score = 0;
      for (const t of textTokens) {
        const nt = normalizeTextForSearch(t);
        if (!nt) continue;
        const nct = normalizeTextCompactJs(nt);
        const inBag = bag.includes(nt) || (nct && bagC.includes(nct));
        if (!inBag) continue;
        const inAlt = altN.includes(nt) || (nct && altC.includes(nct));
        const inItem = itemN.includes(nt) || (nct && itemC.includes(nct));
        if (inAlt) score += 5;
        else if (inItem) score += 2;
      }
      let yb = 0;
      if (years.length) {
        if (years.some(y => yearInRangeTextJs(y, alt))) yb += 3;
        else if (years.some(y => yearInRangeTextJs(y, item))) yb += 1;
      }
      const detail = Math.min(alt.length, 5000);
      return [score + yb, yb, detail, item];
    }

    function scoreTupleBetterJs(a, b) {
      for (let i = 0; i < 4; i++) {
        const x = a[i];
        const y = b[i];
        if (x !== y) return x > y;
      }
      return false;
    }

    function dedupeRowsByNormalizedOriginalForSearchJs(rowsIn, textTokens, years) {
      const tt = textTokens || [];
      const yy = years || [];
      const winners = {};
      for (const r of rowsIn) {
        const oc = normalizeTextCompactJs(String(r.original_numbers || ''));
        if (!oc) continue;
        if (!winners[oc]) winners[oc] = r;
        else {
          const sa = rowSearchRelevanceScoreJs(r, tt, yy);
          const sb = rowSearchRelevanceScoreJs(winners[oc], tt, yy);
          if (scoreTupleBetterJs(sa, sb)) winners[oc] = r;
        }
      }
      const out = [];
      const emitted = new Set();
      for (const r of rowsIn) {
        const oc = normalizeTextCompactJs(String(r.original_numbers || ''));
        if (!oc) {
          out.push(r);
          continue;
        }
        if (emitted.has(oc)) continue;
        emitted.add(oc);
        out.push(winners[oc]);
      }
      return out;
    }

    function searchOfflineSnapshot(query, numberQuery = '', sizeType = '', sizeWidth = '', sizeDiameter = '', sizeHeight = '') {
      const tokens = tokenizeQueryForSearch(query);
      const numberTokens = tokenizeQueryForSearch(numberQuery);
      const hintData = extractFileSearchHintsJs(query, tokens);
      const fileHints = hintData.fileHints;
      if (!hintData.textTokens.length && !numberTokens.length && !sizeType && !sizeWidth && !sizeDiameter && !sizeHeight && !fileHints.length) {
        return { total_rows: 0, matching_item_names: [], matching_items: [], rows: [] };
      }
      const parsedYears = hintData.textTokens.map(parseQueryYearTokenJs);
      const years = Array.from(new Set(parsedYears.filter(Boolean).flat()));
      const explicitYearOps = hintData.textTokens.some(isExplicitYearOperatorTokenJs);
      const textTokens = hintData.textTokens.filter((_, idx) => !parsedYears[idx]);

      let rows = (localSnapshot.rows || []).filter(r => {
        if (!rowMatchesQueryJs(r, textTokens, years)) return false;
        if (fileHints.length) {
          const sf = normalizeTextForSearch(r.source_file || '');
          if (!fileHints.some(h => fileSearchHintMatchesPathJs(sf, h))) return false;
        }
        const numberSpace = normalizeTextForSearch(`${r.company_number || ''} ${r.original_numbers || ''} ${r.notes || ''}`);
        if (!numberTokens.every(t => numberSpace.includes(normalizeTextForSearch(t)))) return false;
        if (!fileMatchesSizeTypeWithFallbackJs(r, sizeType, sizeWidth, sizeDiameter)) return false;
        if (!rowMatchesSizeFiltersJs(r, sizeType, sizeWidth, sizeDiameter, sizeHeight)) return false;
        return true;
      });
      rows = dedupeRowsByNormalizedOriginalForSearchJs(rows, textTokens, years);
      rows = rows.slice(0, 300);
      rows = rows.map((r, i) => {
        const altRaw = String(r.alternatives || '').trim();
        let matched = extractMatchedAlternativeJs(altRaw, textTokens, years, query);
        if (!altRaw) matched = '';
        return {
          ...r,
          item_number: String(r.company_number || r.item_number || ''),
          alternatives: altRaw,
          size_label: formatSizeLabelDisplayJs(r.source_file, r.size_width, r.size_diameter, r.size_height),
          matched_alternative: matched,
          match_score: yearMatchScoreJs(matched, years),
          row_key: `${r.source_file || ''}|${r.source_sheet || ''}|${r.company_number || r.item_number || ''}|${i}`
        };
      });
      if (explicitYearOps && years.length && rows.length) {
        const maxScore = Math.max(...rows.map(r => Number(r.match_score || 0)));
        if (maxScore > 0) rows = rows.filter(r => Number(r.match_score || 0) === maxScore);
      }
      const names = Array.from(new Set(rows.map(r => r.item_name).filter(Boolean))).slice(0, 100);
      const matchingItems = [];
      const seen = new Set();
      for (const r of rows) {
        const key = `${r.item_name}|${r.matched_alternative || ''}`;
        if (!r.item_name || seen.has(key)) continue;
        seen.add(key);
        matchingItems.push({
          item_name: r.item_name,
          matched_alternative: r.matched_alternative || '',
          size_label: r.size_label || '',
          row_key: r.row_key
        });
        if (matchingItems.length >= 100) break;
      }
      return { total_rows: rows.length, matching_item_names: names, matching_items: matchingItems, rows };
    }

    async function syncSnapshotFromServer() {
      try {
        setSyncInfo('جاري تنزيل نسخة البيانات للهاتف...');
        const res = await fetch('/sync/data', { cache: 'no-store' });
        if (!res.ok) throw new Error('sync-data-failed');
        const data = await res.json();
        await saveLocalSnapshot(data);
      } catch (_) {
        setSyncInfo('فشل تنزيل نسخة الهاتف. تأكد أن السيرفر شغال ثم أعد المحاولة.');
      }
    }

    async function syncFromGoogleDrive() {
      const uploadResult = document.getElementById('uploadResult');
      try {
        uploadResult.innerText = 'جاري المزامنة من Google Drive...';
        const res = await fetch('/google_drive/sync', { method: 'POST' });
        const data = await res.json();
        if (!res.ok) throw new Error(data.detail || 'google-drive-sync-failed');

        const unsupportedCount = Array.isArray(data.unsupported_files) ? data.unsupported_files.length : 0;
        const failedCount = Array.isArray(data.failed_files) ? data.failed_files.length : 0;
        uploadResult.innerText =
          `اكتملت مزامنة Google Drive: تم فحص ${Number(data.scanned_files || 0)} ملف، معالجة ${Number(data.files_count || 0)}، تخطي ${Number(data.skipped_files || 0)}، إدخال ${Number(data.inserted_rows || 0)} صف، غير مدعوم ${unsupportedCount}، فشل ${failedCount}.`;

        await loadStats();
        await searchNow();
      } catch (err) {
        uploadResult.innerText = `فشلت مزامنة Google Drive: ${(err && err.message) ? err.message : 'error'}`;
      }
    }

    async function autoSyncGoogleDriveIfServerEmpty() {
      try {
        const doneKey = 'gdrive_auto_sync_done';
        if (sessionStorage.getItem(doneKey) === '1') return;
        const res = await fetch('/stats', { cache: 'no-store' });
        if (!res.ok) return;
        const stats = await res.json();
        if (Number(stats.total_rows || 0) > 0) {
          sessionStorage.setItem(doneKey, '1');
          return;
        }
        // If server is empty, try to restore from Drive automatically.
        await syncFromGoogleDrive();
        sessionStorage.setItem(doneKey, '1');
      } catch (_) {}
    }

    function groupRowsBySourceFile(rows) {
      const groups = new Map();
      for (const row of rows || []) {
        const fileName = String(row.source_file || 'restored-data.xlsx').trim() || 'restored-data.xlsx';
        if (!groups.has(fileName)) groups.set(fileName, []);
        groups.get(fileName).push({
          item_name: String(row.item_name || ''),
          item_number: String(row.item_number || ''),
          company_number: String(row.company_number || ''),
          original_numbers: String(row.original_numbers || ''),
          notes: String(row.notes || ''),
          alternatives: String(row.alternatives || ''),
          size_width: String(row.size_width || ''),
          size_diameter: String(row.size_diameter || ''),
          size_height: String(row.size_height || ''),
          source_sheet: String(row.source_sheet || 'Sheet1')
        });
      }
      return groups;
    }

    async function restoreServerFromLocalSnapshotIfEmpty() {
      try {
        if (!localSnapshot || !Array.isArray(localSnapshot.rows) || !localSnapshot.rows.length) return;

        const statsRes = await fetch('/stats', { cache: 'no-store' });
        if (!statsRes.ok) return;
        const stats = await statsRes.json();
        if (Number(stats.total_rows || 0) > 0) return;

        setSyncInfo('تم اكتشاف سيرفر بدون بيانات. جاري استرجاع الأصناف من النسخة المحلية...');
        const groups = groupRowsBySourceFile(localSnapshot.rows);
        let restoredRows = 0;

        for (const [fileName, rows] of groups.entries()) {
          const pseudoHash = await sha256Hex(new TextEncoder().encode(`${localSnapshot.version || ''}:${fileName}:${rows.length}`));
          const startRes = await fetch('/upload_rows/start', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ file_name: fileName, content_hash: pseudoHash, file_size: rows.length })
          });
          if (!startRes.ok) continue;
          const startData = await startRes.json();
          if (startData.skip) continue;

          const chunkSize = 1200;
          for (let i = 0; i < rows.length; i += chunkSize) {
            const chunk = rows.slice(i, i + chunkSize);
            const cRes = await fetch('/upload_rows/chunk', {
              method: 'POST',
              headers: { 'Content-Type': 'application/json' },
              body: JSON.stringify({ file_name: fileName, rows: chunk })
            });
            if (!cRes.ok) continue;
          }

          const finishRes = await fetch('/upload_rows/finish', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ file_name: fileName, content_hash: pseudoHash, file_size: rows.length })
          });
          if (finishRes.ok) restoredRows += rows.length;
        }

        setSyncInfo(`اكتمل استرجاع البيانات من النسخة المحلية: ${restoredRows} صف.`);
        await loadStats();
      } catch (_) {
        setSyncInfo('تعذر الاسترجاع التلقائي من النسخة المحلية.');
      }
    }

    async function checkSnapshotUpdate(showNoUpdate, askUser = false) {
      try {
        const res = await fetch('/sync/meta', { cache: 'no-store' });
        if (!res.ok) throw new Error('sync-meta-failed');
        const meta = await res.json();  
        if (!localSnapshot.version) {
          setSyncInfo(`متاح تحميل نسخة بيانات (${meta.total_rows} صف).`);
          if (askUser) {
            await syncSnapshotFromServer();
          }
          return;
        }
        if (meta.version !== localSnapshot.version) {
          setSyncInfo('يوجد تحديث جديد للبيانات. اضغط "فحص تحديث البيانات" أو "مزامنة للهاتف".');
          if (askUser && meta.version !== lastNotifiedVersion) {
            await syncSnapshotFromServer();
            lastNotifiedVersion = '';
          } else if (askUser) {
            lastNotifiedVersion = meta.version;
          }
        } else if (showNoUpdate) {
          lastNotifiedVersion = '';
          setSyncInfo('النسخة المحلية محدثة.');
        }
      } catch (_) {
        if (showNoUpdate) setSyncInfo('تعذر فحص التحديث الآن. قد تكون غير متصل بالسيرفر.');
      }
    }

    async function uploadFiles() {
      const selectedFiles = Array.from(selectedFilesMap.values());
      if (!selectedFiles.length) {
        document.getElementById('uploadResult').innerText = 'اختار ملفات Excel أولاً';
        return;
      }

      const uploadBtn = document.getElementById('uploadBtn');
      uploadBtn.disabled = true;
      let done = 0;
      let processed = 0;
      let skipped = 0;
      let inserted = 0;
      let insertedItems = 0;
      const failed = [];
      const tooLarge = [];
      const isVercel = window.location.hostname.includes('vercel.app');
      const vercelMaxBytes = 4_200_000;

      document.getElementById('uploadResult').innerText = 'جاري الرفع والمعالجة...';
      setUploadProgress(0, `بدء الرفع... 0/${selectedFiles.length}`);

      for (const file of selectedFiles) {
        try {
          let data = null;
          const ext = String(file.name || '').toLowerCase().split('.').pop() || '';
          const forceChunked = (ext === 'xlsb' || ext === 'xls');
          if ((isVercel && file.size > vercelMaxBytes) || forceChunked) {
            tooLarge.push(file.name);
            setUploadProgress(
              Math.round((done / selectedFiles.length) * 100),
              forceChunked ? `جاري معالجة ملف بصيغة قديمة: ${file.name}` : `جاري معالجة ملف كبير: ${file.name}`
            );
            data = await uploadLargeFileChunked(file);
          } else {
            const formData = new FormData();
            formData.append('files', file, sourceKeyForFile(file));
            const res = await fetch('/upload', { method: 'POST', body: formData });
            data = await readJsonSafe(res);
            if (!res.ok) {
              if (res.status === 413 || ext === 'xls' || ext === 'xlsb') {
                // fallback to chunked mode automatically
                data = await uploadLargeFileChunked(file);
              } else {
                throw new Error(data.detail || 'upload-failed');
              }
            }
          }
          processed += Number(data.files_count || 0);
          skipped += Number(data.skipped_files || 0);
          inserted += Number(data.inserted_rows || 0);
          insertedItems += Number(data.inserted_items || 0);
        } catch (err) {
          failed.push(`${file.name} (${(err && err.message) ? err.message : 'upload-failed'})`);
        }

        done += 1;
        const percent = Math.round((done / selectedFiles.length) * 100);
        setUploadProgress(percent, `جاري الرفع: ${done}/${selectedFiles.length} - الملف الحالي: ${file.name}`);
      }

      uploadBtn.disabled = false;
      const failText = failed.length ? ` | فشل ${failed.length} ملف` : '';
      const largeText = tooLarge.length ? ` | ملفات تمت معالجتها بنمط مجزأ: ${tooLarge.length}` : '';
      document.getElementById('uploadResult').innerText =
        `اكتمل الرفع ${Math.round((done / selectedFiles.length) * 100)}%: تم معالجة ${processed} ملفات، تخطي ${skipped} ملفات غير معدلة، وإدخال ${inserted} صف، وإضافة ${insertedItems} صنف${largeText}${failText}${failed.length ? ` | أمثلة فشل: ${failed.slice(0, 3).join(' | ')}` : ''}.`;
      await loadStats();
      await searchNow();
    }

    async function searchNow(mode = 'auto') {
      const effectiveMode = mode === 'auto' ? activeSearchTab : mode;

      const qRaw = document.getElementById('queryInput').value.trim();
      const qNumbersRaw = document.getElementById('numberQueryInput').value.trim();
      const sizeTypeRaw = document.getElementById('sizeTypeInput').value.trim();
      const sizeWidthRaw = document.getElementById('sizeWidthInput').value.trim();
      const sizeDiameterRaw = document.getElementById('sizeDiameterInput').value.trim();
      const sizeHeightRaw = document.getElementById('sizeHeightInput').value.trim();

      const q = effectiveMode === 'main' ? qRaw : '';
      const qNumbers = effectiveMode === 'main' ? qNumbersRaw : '';
      const sizeType = effectiveMode === 'size' ? sizeTypeRaw : '';
      const sizeWidth = effectiveMode === 'size' ? sizeWidthRaw : '';
      const sizeDiameter = effectiveMode === 'size' ? sizeDiameterRaw : '';
      const sizeHeight = effectiveMode === 'size' ? sizeHeightRaw : '';

      localStorage.setItem('lastSearchQuery', q);
      localStorage.setItem('lastNumberQuery', qNumbers);
      localStorage.setItem('lastSizeType', sizeType);
      localStorage.setItem('lastSizeWidth', sizeWidth);
      localStorage.setItem('lastSizeDiameter', sizeDiameter);
      localStorage.setItem('lastSizeHeight', sizeHeight);

      if (!q && !qNumbers && !sizeType && !sizeWidth && !sizeDiameter && !sizeHeight) {
        const namesDiv = document.getElementById('names');
        namesDiv.innerHTML = '';
        const table = document.getElementById('resultsTable');
        table.style.display = 'none';
        document.getElementById('stats').innerText = effectiveMode === 'main'
          ? 'لا توجد نتائج. اكتب كلمة بحث.'
          : 'لا توجد نتائج. اكتب القياسات أو اختر النوع.';
        return;
      }

      const params = new URLSearchParams({
        q: q,
        q_numbers: qNumbers,
        size_type: sizeType,
        size_width: sizeWidth,
        size_diameter: sizeDiameter,
        size_height: sizeHeight
      });

      let data = null;
      let sourceLabel = 'الخادم';
      try {
        const ctrl = new AbortController();
        const to = setTimeout(() => ctrl.abort(), 90000);
        const res = await fetch(`/search?${params.toString()}`, {
          method: 'GET',
          cache: 'no-store',
          signal: ctrl.signal,
          headers: { Accept: 'application/json' }
        });
        clearTimeout(to);
        const parsed = await readJsonSafe(res);
        if (!res.ok) throw new Error(parsed.detail || `search-http-${res.status}`);
        if (!parsed || typeof parsed !== 'object') throw new Error('search-bad-json');
        data = parsed;
      } catch (_) {
        data = searchOfflineSnapshot(q, qNumbers, sizeType, sizeWidth, sizeDiameter, sizeHeight);
        sourceLabel = 'النسخة المحلية';

        // Mobile fallback: if local data is empty but internet exists,
        // pull a snapshot once and retry to avoid "no data" confusion.
        if (!(localSnapshot.rows || []).length && navigator.onLine) {
          try {
            await syncSnapshotFromServer();
            data = searchOfflineSnapshot(q, qNumbers, sizeType, sizeWidth, sizeDiameter, sizeHeight);
            if (data.total_rows > 0) {
              sourceLabel = 'النسخة المحلية (بعد تنزيل البيانات)';
            } else {
              const ctrl2 = new AbortController();
              const to2 = setTimeout(() => ctrl2.abort(), 90000);
              const retry = await fetch(`/search?${params.toString()}`, {
                method: 'GET',
                cache: 'no-store',
                signal: ctrl2.signal,
                headers: { Accept: 'application/json' }
              });
              clearTimeout(to2);
              const parsed2 = await readJsonSafe(retry);
              if (retry.ok && parsed2 && typeof parsed2 === 'object') {
                data = parsed2;
                sourceLabel = 'الخادم';
              }
            }
          } catch (_) {}
        }
      }

      const safeRows = Array.isArray(data && data.rows) ? data.rows : [];
      const safeTotal = Number((data && data.total_rows) || safeRows.length || 0);
      document.getElementById('stats').innerText = `عدد النتائج: ${safeTotal} | المصدر: ${sourceLabel}`;
      if (!safeTotal && sourceLabel.includes('النسخة المحلية') && !(localSnapshot.rows || []).length) {
        setSyncInfo('لا توجد بيانات محفوظة محلياً بعد. استخدم "مزامنة للهاتف (Offline)" مرة واحدة وأعد البحث.');
      }

      const namesDiv = document.getElementById('names');
      namesDiv.innerHTML = '';
      const quickItemsRaw = Array.isArray(data && data.matching_items)
        ? data.matching_items
        : (Array.isArray(data && data.matching_item_names) ? data.matching_item_names.map(name => ({
            item_name: name,
            matched_alternative: '',
            size_label: '',
            row_key: name
          })) : []);
      const quickItems = quickItemsRaw.map((entry) => {
        if (entry && typeof entry === 'object') return entry;
        return {
          item_name: String(entry || ''),
          matched_alternative: '',
          size_label: '',
          row_key: String(entry || '')
        };
      }).filter((entry) => String(entry.item_name || '').trim());

      /** إزالة لاحقة مثل « بربيش ... ماء.xlsx» إذا وُسِم اسم الصنف باسم ملف في البيانات */
      function stripTrailingSpreadsheetFileFromDisplayName(label) {
        const s = String(label || '').trim();
        if (!s) return s;
        const re = /\s+([^.]+)\.(xlsx|xlsm|xlsb|xls)\s*$/i;
        const m = re.exec(s);
        if (!m || m.index === undefined) return s;
        const body = String(m[1] || '').trim();
        if (!/[\u0600-\u06FF]/.test(body)) return s;
        return s.slice(0, m.index).trim();
      }

      quickItems.forEach(entry => {
        const btn = document.createElement('button');
        btn.className = 'pill pill-btn';
        btn.type = 'button';
        const quickName = document.createElement('span');
        quickName.className = 'quick-name';
        quickName.textContent = stripTrailingSpreadsheetFileFromDisplayName(entry.item_name || '');
        btn.appendChild(quickName);
        if (entry.matched_alternative) {
          const quickAlt = document.createElement('span');
          quickAlt.className = 'quick-alt';
          quickAlt.textContent = entry.matched_alternative;
          btn.appendChild(quickAlt);
        }
        if (entry.size_label) {
          const quickSize = document.createElement('span');
          quickSize.className = 'quick-size';
          quickSize.textContent = entry.size_label;
          btn.appendChild(quickSize);
        }
        btn.onclick = () => focusRowByKey(entry.row_key);
        namesDiv.appendChild(btn);
      });

      const table = document.getElementById('resultsTable');
      const body = table.querySelector('tbody');
      body.innerHTML = '';
      safeRows.forEach(r => {
        const sizeLabel = String(r.size_label || '');
        const nameForUi = stripTrailingSpreadsheetFileFromDisplayName(r.item_name || '');
        const nameCell = `${escapeHtml(nameForUi)}${sizeLabel ? `<div class="quick-size">${escapeHtml(sizeLabel)}</div>` : ''}`;
        const itemNumberDisplay = String(r.company_number || r.item_number || '');
        const fileName = String(r.source_file || '').trim();
        const tr = document.createElement('tr');
        tr.dataset.rowKey = String(r.row_key || '');
        tr.innerHTML = `<td>${escapeHtml(fileName)}</td>
                        <td>${nameCell}</td>
                        <td>${escapeHtml(itemNumberDisplay)}</td>
                        <td>${escapeHtml(String(r.alternatives ?? ''))}</td>`;
        body.appendChild(tr);
      });
      table.style.display = safeRows.length ? 'table' : 'none';
    }

    function initAutoSearch() {
      const queryInput = document.getElementById('queryInput');
      const savedQuery = localStorage.getItem('lastSearchQuery') || '';
      queryInput.value = savedQuery;
      document.getElementById('numberQueryInput').value = localStorage.getItem('lastNumberQuery') || '';
      document.getElementById('sizeTypeInput').value = localStorage.getItem('lastSizeType') || '';
      document.getElementById('sizeWidthInput').value = localStorage.getItem('lastSizeWidth') || '';
      document.getElementById('sizeDiameterInput').value = localStorage.getItem('lastSizeDiameter') || '';
      document.getElementById('sizeHeightInput').value = localStorage.getItem('lastSizeHeight') || '';
      setSearchTab(localStorage.getItem('activeSearchTab') || 'main');
      updateSizeInputsLayout();

      queryInput.addEventListener('input', () => {
        clearTimeout(searchDebounceTimer);
        searchDebounceTimer = setTimeout(() => {
          searchNow('main');
        }, 250);
      });
      document.getElementById('numberQueryInput').addEventListener('input', () => {
        clearTimeout(searchDebounceTimer);
        searchDebounceTimer = setTimeout(() => {
          searchNow('main');
        }, 250);
      });
      document.getElementById('sizeTypeInput').addEventListener('change', () => {
        updateSizeInputsLayout();
        searchNow('size');
      });
      document.getElementById('sizeWidthInput').addEventListener('input', () => {
        clearTimeout(searchDebounceTimer);
        searchDebounceTimer = setTimeout(() => searchNow('size'), 250);
      });
      document.getElementById('sizeDiameterInput').addEventListener('input', () => {
        clearTimeout(searchDebounceTimer);
        searchDebounceTimer = setTimeout(() => searchNow('size'), 250);
      });
      document.getElementById('sizeHeightInput').addEventListener('input', () => {
        clearTimeout(searchDebounceTimer);
        searchDebounceTimer = setTimeout(() => searchNow('size'), 250);
      });

      clearInterval(autoRefreshTimer);
      autoRefreshTimer = setInterval(() => {
        loadStats();
      }, 10000);
    }

    function focusRowByKey(rowKey) {
      const rows = Array.from(document.querySelectorAll('#resultsTable tbody tr'));
      const target = rows.find(r => r.dataset.rowKey === String(rowKey || ''));
      if (!target) return;

      rows.forEach(r => r.classList.remove('row-focus'));
      target.classList.add('row-focus');
      target.scrollIntoView({ behavior: 'smooth', block: 'center' });
    }

    function escapeHtml(str) {
      return String(str || '')
        .replace(/&/g, '&amp;')
        .replace(/</g, '&lt;')
        .replace(/>/g, '&gt;')
        .replace(/"/g, '&quot;')
        .replace(/'/g, '&#039;');
    }

    async function loadStats() {
      try {
        const res = await fetch('/stats', { cache: 'no-store' });
        const data = await res.json();
        const backendKey = String(data.db_backend || 'sqlite').toLowerCase();
        const backend = backendKey === 'supabase'
          ? 'Supabase/Postgres'
          : backendKey === 'neon'
            ? 'Neon/Postgres'
            : backendKey === 'postgres'
              ? 'Postgres'
              : 'SQLite';
        document.getElementById('stats').innerText =
          `إجمالي الصفوف: ${data.total_rows} | عدد الأصناف: ${data.total_items || 0} | عدد الملفات: ${data.total_files} | المصدر: الخادم | قاعدة البيانات: ${backend}`;
      } catch (_) {
        document.getElementById('stats').innerText =
          `إجمالي الصفوف (محلي): ${localSnapshot.total_rows || (localSnapshot.rows || []).length}`;
      }
    }
    const dropZone = document.getElementById('dropZone');
    if (dropZone) {
      dropZone.addEventListener('dragover', (e) => {
        e.preventDefault();
        dropZone.classList.add('active');
      });
      dropZone.addEventListener('dragleave', () => {
        dropZone.classList.remove('active');
      });
      dropZone.addEventListener('drop', (e) => {
        handleDrop(e);
      });
    }
    updateSelectionInfo();
    setUploadProgress(0, '');
    registerServiceWorker();
    (async () => {
      await ensureLatestClientVersion();
      await loadLocalSnapshot();
      initAutoSearch();
      loadStats();
      searchNow();
      restoreServerFromLocalSnapshotIfEmpty();
      autoSyncGoogleDriveIfServerEmpty();
    })();
  </script>
</body>
</html>
"""


@app.post("/upload")
async def upload(files: list[UploadFile] = File(...)) -> dict:
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    inserted_total = 0
    inserted_items_total = 0
    processed_files = 0
    skipped_files = 0

    for file in files:
        lower = file.filename.lower() if file.filename else ""
        if not lower.endswith((".xlsx", ".xls", ".xlsm")):
            continue

        content = await file.read()
        content_hash, file_size = file_fingerprint(content)
        if is_same_uploaded_file(file.filename, content_hash, file_size):
            skipped_files += 1
            continue
        try:
            inserted_rows, inserted_items = process_excel_file(content, file.filename, content_hash, file_size)
            inserted_total += inserted_rows
            inserted_items_total += inserted_items
            processed_files += 1
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Could not process file '{file.filename}': {exc}",
            ) from exc

    return {
        "files_count": processed_files,
        "inserted_rows": inserted_total,
        "inserted_items": inserted_items_total,
        "skipped_files": skipped_files,
    }


@app.post("/google_drive/sync")
def google_drive_sync() -> dict:
    folder_id = (os.getenv("GDRIVE_FOLDER_ID") or "").strip()
    if not folder_id:
        raise HTTPException(status_code=400, detail="Missing GDRIVE_FOLDER_ID")

    try:
        service = get_gdrive_service()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Google Drive setup error: {exc}") from exc

    files = list_gdrive_excel_files(service, folder_id)
    processed = 0
    skipped = 0
    inserted_rows_total = 0
    inserted_items_total = 0
    failed: list[str] = []
    unsupported: list[str] = []

    for f in files:
        file_name = normalize_source_file_name(f.get("relative_name") or f.get("name") or "")
        if not file_name:
            continue
        if not is_openpyxl_supported_file_name(file_name):
            unsupported.append(file_name)
            continue

        try:
            content = download_gdrive_file_bytes(service, f["id"])
            content_hash, file_size = file_fingerprint(content)
            if is_same_uploaded_file(file_name, content_hash, file_size):
                skipped += 1
                continue
            rows, items = process_excel_file(content, file_name, content_hash, file_size)
            processed += 1
            inserted_rows_total += rows
            inserted_items_total += items
        except Exception as exc:
            failed.append(f"{file_name}: {exc}")

    return {
        "scanned_files": len(files),
        "files_count": processed,
        "skipped_files": skipped,
        "inserted_rows": inserted_rows_total,
        "inserted_items": inserted_items_total,
        "unsupported_files": unsupported[:100],
        "failed_files": failed[:100],
    }


@app.post("/upload_rows/start")
def upload_rows_start(payload: UploadRowsStartIn) -> dict:
    source_key = normalize_source_file_name(payload.file_name)
    if is_same_uploaded_file(source_key, payload.content_hash, payload.file_size):
        return {"skip": True}

    conn = get_db()
    try:
        conn.execute("DELETE FROM products WHERE source_file = ?", (source_key,))
        conn.commit()
    finally:
        conn.close()
    return {"skip": False}


@app.post("/upload_rows/chunk")
def upload_rows_chunk(payload: UploadRowsChunkIn) -> dict:
    if not payload.rows:
        return {"inserted_rows": 0}

    source_key = normalize_source_file_name(payload.file_name)
    rows_to_insert = []
    for row in payload.rows:
        if not any(
            [
                row.item_name,
                row.item_number,
                row.company_number,
                row.original_numbers,
                row.notes,
                row.alternatives,
                row.size_width,
                row.size_diameter,
                row.size_height,
            ]
        ):
            continue
        rows_to_insert.append(
            (
                row.item_name.strip(),
                row.item_number.strip(),
                row.company_number.strip(),
                row.original_numbers.strip(),
                row.notes.strip(),
                sanitize_alternatives_vs_original(row.alternatives, row.original_numbers),
                row.size_width.strip(),
                row.size_diameter.strip(),
                row.size_height.strip(),
                source_key,
                (row.source_sheet or "Sheet1").strip(),
            )
        )

    if not rows_to_insert:
        return {"inserted_rows": 0}

    conn = get_db()
    try:
        conn.executemany(
            """
            INSERT INTO products (
                item_name, item_number, company_number, original_numbers, notes, alternatives,
                size_width, size_diameter, size_height, source_file, source_sheet
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows_to_insert,
        )
        conn.commit()
    finally:
        conn.close()
    return {"inserted_rows": len(rows_to_insert)}


@app.post("/upload_rows/finish")
def upload_rows_finish(payload: UploadRowsFinishIn) -> dict:
    source_key = normalize_source_file_name(payload.file_name)
    conn = get_db()
    try:
        rows_count = conn.execute(
            "SELECT COUNT(*) AS c FROM products WHERE source_file = ?",
            (source_key,),
        ).fetchone()["c"]
    finally:
        conn.close()

    upsert_uploaded_file_meta(source_key, rows_count, payload.content_hash, payload.file_size)
    return {"rows_count": rows_count}


@app.post("/upload_rows/abort")
def upload_rows_abort(payload: UploadRowsAbortIn) -> dict:
    source_key = normalize_source_file_name(payload.file_name)
    conn = get_db()
    try:
        conn.execute("DELETE FROM products WHERE source_file = ?", (source_key,))
        conn.execute("DELETE FROM uploaded_files WHERE file_name = ?", (source_key,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.get("/search")
def search(
    q: str = "",
    q_numbers: str = "",
    size_type: str = "",
    size_width: str = "",
    size_diameter: str = "",
    size_height: str = "",
) -> dict:
    tokens = tokenize_query(q)
    number_tokens = tokenize_query(q_numbers)
    file_hints, tokens = extract_file_search_hints(q, tokens)
    size_active = any([size_type.strip(), size_width.strip(), size_diameter.strip(), size_height.strip()])
    if not tokens and not number_tokens and not size_active and not file_hints:
        return {"total_rows": 0, "matching_item_names": [], "matching_items": [], "rows": []}

    parsed_year_tokens = [parse_query_year_token(t) for t in tokens]
    query_years = sorted({y for ys in parsed_year_tokens if ys for y in ys})
    explicit_year_ops = any(is_explicit_year_operator_token(t) for t in tokens)
    text_tokens = [t for t, ys in zip(tokens, parsed_year_tokens) if ys is None]

    conn = get_db()
    try:
        where_parts = []
        params: list[str] = []

        for token in text_tokens:
            # Digits can appear in Arabic/Persian forms inside Excel text.
            # Keep final matching in Python (normalized) to avoid false misses.
            if any(ch.isdigit() for ch in token):
                continue
            like = f"%{token}%"
            where_parts.append("(lower(item_name) LIKE ? OR lower(alternatives) LIKE ? OR lower(source_file) LIKE ?)")
            params.extend([like, like, like])

        for token in number_tokens:
            if any(ch.isdigit() for ch in token):
                continue
            like = f"%{token}%"
            where_parts.append("(lower(company_number) LIKE ? OR lower(original_numbers) LIKE ? OR lower(notes) LIKE ?)")
            params.extend([like, like, like])

        if file_hints:
            fh_sql, fh_params = file_hints_sql_clause(file_hints)
            where_parts.append(fh_sql)
            params.extend(fh_params)

        # تصفية أولية بالقياسات (أرقام فقط) لتقليل الصفوف قبل فلترة بايثون
        if size_active:
            for raw in (size_width, size_diameter, size_height):
                v = normalize_text(raw or "").replace(",", ".")
                if v and re.fullmatch(r"[\d.]+", v):
                    likev = f"%{v.lower()}%"
                    where_parts.append(
                        "(lower(COALESCE(size_width,'')) LIKE ? OR "
                        "lower(COALESCE(size_diameter,'')) LIKE ? OR "
                        "lower(COALESCE(size_height,'')) LIKE ?)"
                    )
                    params.extend([likev, likev, likev])

        # Size/type filtering is applied in Python for better tolerance.

        where_clause = " AND ".join(where_parts) if where_parts else "1=1"
        candidate_rows = conn.execute(
            f"""
            SELECT item_name, item_number, company_number, original_numbers, notes, alternatives,
                   size_width, size_diameter, size_height, source_file, source_sheet
            FROM products
            WHERE {where_clause}
            ORDER BY item_name
            LIMIT 40000
            """,
            params,
        ).fetchall()

        filtered_rows = [
            r
            for r in candidate_rows
            if row_matches_query(r, text_tokens, query_years)
            and row_matches_number_query(r, number_tokens)
            and (
                not file_hints
                or any(file_search_hint_matches_path(r["source_file"] or "", h) for h in file_hints)
            )
            and file_matches_size_type_with_fallback(r, r["source_file"] or "", size_type, size_width, size_diameter)
            and row_matches_size_filters(size_type, r, size_width, size_diameter, size_height)
        ]
        filtered_rows = dedupe_rows_by_normalized_original_for_search(
            filtered_rows, text_tokens, query_years
        )
        rows = filtered_rows[:500]
    finally:
        conn.close()

    prepared_rows = []
    quick_items = []
    seen_quick = set()
    for index, row in enumerate(rows):
        row_dict = dict(row)
        matched_alt = extract_matched_alternative(
            row_dict.get("alternatives", ""),
            text_tokens,
            query_years,
            q,
        )
        if not (row_dict.get("alternatives") or "").strip():
            matched_alt = ""
        row_key = f"{row_dict.get('source_file', '')}|{row_dict.get('source_sheet', '')}|{row_dict.get('company_number') or row_dict.get('item_number', '')}|{index}"
        size_label = build_size_label_display(
            row_dict.get("source_file", "") or "",
            row_dict.get("size_width") or "",
            row_dict.get("size_diameter") or "",
            row_dict.get("size_height") or "",
        )
        item_number_display = (row_dict.get("company_number") or row_dict.get("item_number") or "").strip()
        row_dict["item_number"] = item_number_display
        row_dict["size_label"] = size_label
        row_dict["matched_alternative"] = matched_alt
        row_dict["alternatives"] = (row_dict.get("alternatives") or "").strip()
        row_dict["match_score"] = year_match_score(matched_alt, query_years)
        row_dict["row_key"] = row_key
        prepared_rows.append(row_dict)

        quick_key = (row_dict.get("item_name", ""), matched_alt)
        if quick_key not in seen_quick and row_dict.get("item_name", ""):
            seen_quick.add(quick_key)
            quick_items.append(
                {
                    "item_name": row_dict["item_name"],
                    "matched_alternative": matched_alt,
                    "size_label": row_dict.get("size_label", ""),
                    "row_key": row_key,
                }
            )
        if len(quick_items) >= 100:
            break

    # Keep "most precise only" behavior for explicit year operators (+YY, YY-, YY-YY),
    # but keep all valid matches for plain year tokens like "20".
    if explicit_year_ops and query_years and prepared_rows:
        max_score = max(r.get("match_score", 0) for r in prepared_rows)
        if max_score > 0:
            prepared_rows = [r for r in prepared_rows if r.get("match_score", 0) == max_score]

            # Rebuild quick items after precision filtering.
            quick_items = []
            seen_quick = set()
            for r in prepared_rows:
                quick_key = (r.get("item_name", ""), r.get("matched_alternative", ""))
                if quick_key in seen_quick or not r.get("item_name"):
                    continue
                seen_quick.add(quick_key)
                quick_items.append(
                    {
                        "item_name": r["item_name"],
                        "matched_alternative": r.get("matched_alternative", ""),
                        "size_label": r.get("size_label", ""),
                        "row_key": r.get("row_key", ""),
                    }
                )
                if len(quick_items) >= 100:
                    break

    names = sorted({r["item_name"] for r in prepared_rows if r["item_name"]})[:100]
    return {
        "total_rows": len(prepared_rows),
        "matching_item_names": names,
        "matching_items": quick_items,
        "rows": prepared_rows,
    }


@app.get("/stats")
def stats() -> dict:
    conn = get_db()
    try:
        total_rows = conn.execute("SELECT COUNT(*) AS c FROM products").fetchone()["c"]
        total_files = conn.execute("SELECT COUNT(*) AS c FROM uploaded_files").fetchone()["c"]
        total_items = conn.execute(
            "SELECT COUNT(DISTINCT item_name) AS c FROM products WHERE COALESCE(TRIM(item_name), '') <> ''"
        ).fetchone()["c"]
    finally:
        conn.close()
    return {
        "total_rows": total_rows,
        "total_files": total_files,
        "total_items": total_items,
        "db_backend": DB_BACKEND,
    }


@app.get("/sync/meta")
def sync_meta() -> dict:
    return get_sync_meta()


@app.get("/sync/data")
def sync_data() -> dict:
    meta = get_sync_meta()
    return {
        "version": meta["version"],
        "updated_at": meta["updated_at"],
        "total_rows": meta["total_rows"],
        "rows": get_sync_rows(),
    }


@app.get("/manifest.webmanifest")
def manifest() -> HTMLResponse:
    payload = {
        "name": "بحث البدائل",
        "short_name": "بدائل",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f6f7fb",
        "theme_color": "#0b66ff",
        "lang": "ar",
    }
    return HTMLResponse(content=json.dumps(payload, ensure_ascii=False), media_type="application/manifest+json")


@app.get("/sw.js")
def service_worker() -> HTMLResponse:
    js = f"""
const CACHE_NAME = 'badail-shell-{APP_VERSION}';
const APP_SHELL = ['/', '/manifest.webmanifest'];

self.addEventListener('install', (event) => {{
  event.waitUntil(caches.open(CACHE_NAME).then((cache) => cache.addAll(APP_SHELL)));
  self.skipWaiting();
}});

self.addEventListener('activate', (event) => {{
  event.waitUntil(
    caches.keys().then((keys) => Promise.all(keys.filter((k) => k.startsWith('badail-shell-') && k !== CACHE_NAME).map((k) => caches.delete(k))))
  );
  self.clients.claim();
}});

self.addEventListener('fetch', (event) => {{
  const req = event.request;
  if (req.method !== 'GET') return;

  if (req.mode === 'navigate') {{
    event.respondWith(
      fetch(req).then((res) => {{
        const copy = res.clone();
        caches.open(CACHE_NAME).then((cache) => cache.put('/', copy));
        return res;
      }}).catch(() => caches.match('/'))
    );
    return;
  }}

  if (req.url.includes('/sync/') || req.url.includes('/search') || req.url.includes('/stats') || req.url.includes('/upload')) {{
    event.respondWith(fetch(req).catch(() => caches.match(req)));
    return;
  }}

  event.respondWith(
    caches.match(req).then((cached) => cached || fetch(req).then((res) => {{
      const copy = res.clone();
      caches.open(CACHE_NAME).then((cache) => cache.put(req, copy));
      return res;
    }}).catch(() => caches.match('/')))
  );
}});
"""
    return HTMLResponse(content=js, media_type="application/javascript")


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "app_version": APP_VERSION, "db_backend": DB_BACKEND}
